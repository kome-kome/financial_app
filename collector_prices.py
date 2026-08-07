"""株価収集（stooq / J-Quants / Yahoo Finance）とマクロ指標収集。"""
import bisect
import calendar
import csv
import io
import zipfile
import asyncio
from collections import defaultdict
from datetime import date, timedelta
from typing import Optional, Callable
from urllib.parse import quote as urlquote  # fetch_yahoo_history のローカル変数 quote と衝突回避

import httpx
import openpyxl
import pandas as pd
from sqlalchemy import func as sqla_func
from sqlalchemy.exc import SQLAlchemyError

from database import (
    SessionLocal, Company, FinancialRecord, MacroData,
    XbrlRawDocument, upsert_company, upsert_financial,
    upsert_xbrl_raw, pack_elements, unpack_elements,
    build_xbrl_map,
    StockPriceDaily, StockPriceWeekly, DAILY_WINDOW_DAYS,
    record_prices_batch, trim_daily, latest_prices,
    upsert_macro_batch, sync_active_status,
)

from collector_utils import *


def _stooq_float(s: str) -> float | None:
    """stooq CSV セルを float 化。パース不能なら None（欠損許容経路用）。"""
    try:
        return float(s)
    except ValueError:
        return None


def _parse_stooq_csv(text: str, *, strict: bool) -> list:
    """stooq 日次 OHLCV CSV（"Date,Open,High,Low,Close,Volume"）をパースする。
    close は両経路とも必須（パース不能行はスキップ）。
    strict=True : open/high/low/volume も float 必須で、不能なら行ごとスキップ（個別銘柄経路）。
    strict=False: open/high/low/volume は None 許容（マクロ経路）。
    """
    rows = []
    lines = text.strip().splitlines()
    if len(lines) < 2:
        return []
    for line in lines[1:]:   # ヘッダー行をスキップ
        parts = line.split(",")
        if len(parts) < 5:
            continue
        try:
            close = float(parts[4])
        except ValueError:
            continue
        if strict:
            try:
                row = {
                    "trade_date": parts[0],          # "YYYY-MM-DD"
                    "open":       float(parts[1]),
                    "high":       float(parts[2]),
                    "low":        float(parts[3]),
                    "close":      close,
                    "volume":     float(parts[5]) if len(parts) > 5 else None,
                }
            except ValueError:
                continue
        else:
            row = {
                "trade_date": parts[0],
                "open":       _stooq_float(parts[1]),
                "high":       _stooq_float(parts[2]),
                "low":        _stooq_float(parts[3]),
                "close":      close,
                "volume":     _stooq_float(parts[5]) if len(parts) > 5 else None,
            }
        rows.append(row)
    return rows


async def _fetch_stooq_ohlcv(
    session: httpx.AsyncClient,
    ticker: str,
    date_from: str,   # "YYYYMMDD"
    date_to: str,     # "YYYYMMDD"
    *,
    strict: bool,
    log_label: str,
) -> list:
    """stooq 日次 OHLCV CSV を取得・パースする単一実装。
    ticker 組み立ては呼び出し側が行う（個別銘柄は `.jp` 付与・マクロはそのまま）。"""
    url = f"https://stooq.com/q/d/l/?s={ticker}&d1={date_from}&d2={date_to}&i=d"
    try:
        r = await session.get(url, timeout=30)
        r.raise_for_status()
        text = r.text
    except Exception as e:
        log.debug(f"{log_label}: {e}")
        return []
    return _parse_stooq_csv(text, strict=strict)


async def fetch_stock_history_stooq(
    session: httpx.AsyncClient,
    sec_code: str,
    date_from: str,   # "YYYYMMDD"
    date_to: str,     # "YYYYMMDD"
) -> list:
    """stooq 日次 OHLCV を取得して [{trade_date, open, high, low, close, volume}] で返す（個別銘柄・`.jp` 付与）。"""
    return await _fetch_stooq_ohlcv(
        session, f"{sec_code}.jp", date_from, date_to,
        strict=True, log_label=f"stooq履歴取得失敗 {sec_code}",
    )


async def _price_collection_driver(db, batch_gen) -> tuple[bool, int]:
    """
    Provider-agnostic driver: iterates batch_gen, saves each batch via
    record_prices_batch, calls trim_daily when done.

    batch_gen yields list[dict] of price records, or None as a cancellation
    sentinel (triggers early return with cancelled=True).
    Returns (cancelled, total_inserted).
    """
    total = 0
    async for batch in batch_gen:
        if batch is None:   # cancellation sentinel from generator
            db.commit()
            return True, total
        if batch:
            try:
                total += record_prices_batch(db, batch, trim=False)
            except Exception as e:
                log.warning("株価バッチ保存失敗: %s", e)
                db.rollback()  # aborted transaction をリセット。次バッチ・trim_daily を救済
    trim_daily(db)
    return False, total


async def collect_stock_price_history(
    db,
    years_back: int = 3,
    max_companies: Optional[int] = None,
    on_progress: Optional[Callable] = None,
    cancel_check: Optional[Callable] = None,
    skip_existing: bool = True,
    backfill: bool = False,
) -> dict:
    """全企業（sec_code 保有）の日次 OHLCV を stooq から取得して DB に保存する。
    skip_existing=True: DB の最新 trade_date から翌日以降のみ取得（差分収集）。
    backfill=True かつ skip_existing=True: 前方差分に加えて後方欠損（years_back 起点→最古レコード前日）も補完。
    プロバイダー固有ロジック（stooq 並行フェッチ）を _stooq_batch_gen に分離し、
    _price_collection_driver の共通フレームで DB 保存・trim を一元管理する。
    """

    today     = date.today()
    date_from = date(today.year - years_back, today.month, today.day)
    d1 = date_from.strftime("%Y%m%d")
    d2 = today.strftime("%Y%m%d")
    date_from_str = date_from.strftime("%Y-%m-%d")
    yesterday = (today - timedelta(days=1)).strftime("%Y-%m-%d")

    companies = (
        db.query(Company.edinet_code, Company.sec_code, Company.name)
        .filter(Company.sec_code.isnot(None), Company.sec_code != "")
        .all()
    )
    if max_companies:
        companies = companies[:max_companies]

    # 差分収集: 企業ごとの min/max の trade_date を一括取得（ループ外で1回のみ）
    minmax_dates: dict = {}
    latest_dates: dict = {}
    if skip_existing:
        # 差分判定は全履歴を持つ weekly の min/max を基準にする（daily は直近窓のみのため）
        if backfill:
            minmax_dates = {
                row.edinet_code: (row.min_date, row.max_date)
                for row in db.query(
                    StockPriceWeekly.edinet_code,
                    sqla_func.min(StockPriceWeekly.trade_date).label("min_date"),
                    sqla_func.max(StockPriceWeekly.trade_date).label("max_date"),
                ).group_by(StockPriceWeekly.edinet_code).all()
            }
        else:
            latest_dates = dict(
                db.query(StockPriceWeekly.edinet_code, sqla_func.max(StockPriceWeekly.trade_date))
                .group_by(StockPriceWeekly.edinet_code)
                .all()
            )

    total = len(companies)
    skipped_total  = 0

    # 差分収集: スキップ判定を事前に行い、取得対象だけリストアップ
    # to_fetch: (edinet_code, sec_code, name, d1_co, d2_co) のリスト
    to_fetch = []
    for edinet_code, sec_code, name in companies:
        if skip_existing and backfill:
            entry = minmax_dates.get(edinet_code)
            if entry is None:
                to_fetch.append((edinet_code, sec_code, name, d1, d2))
            else:
                min_date, max_date = entry
                added = False
                if max_date < yesterday:
                    d1_fwd = (date.fromisoformat(max_date) + timedelta(days=1)).strftime("%Y%m%d")
                    to_fetch.append((edinet_code, sec_code, name, d1_fwd, d2))
                    added = True
                if min_date > date_from_str:
                    min_dt = date.fromisoformat(min_date)
                    if min_dt > date_from:
                        d2_bwd = (min_dt - timedelta(days=1)).strftime("%Y%m%d")
                        to_fetch.append((edinet_code, sec_code, name, d1, d2_bwd))
                        added = True
                if not added:
                    skipped_total += 1
        elif skip_existing:
            latest = latest_dates.get(edinet_code)
            if latest and latest >= yesterday:
                skipped_total += 1
                continue
            d1_company = (date.fromisoformat(latest) + timedelta(days=1)).strftime("%Y%m%d") if latest else d1
            to_fetch.append((edinet_code, sec_code, name, d1_company, d2))
        else:
            to_fetch.append((edinet_code, sec_code, name, d1, d2))

    fetch_total = len(to_fetch)
    progress_total = fetch_total if backfill else total
    if on_progress and skipped_total:
        on_progress(0 if backfill else skipped_total, progress_total,
                    f"[スキップ] {skipped_total}社は補完不要 → {fetch_total}件を取得します" if backfill
                    else f"[スキップ] {skipped_total}社は最新済み → {fetch_total}社を並列取得します")

    async def _stooq_batch_gen(session):
        sem = asyncio.Semaphore(STOOQ_HIST_CONCURRENCY)

        async def _fetch_hist(ec, sc, nm, d1c, d2c):
            async with sem:
                rows = await fetch_stock_history_stooq(session, sc, d1c, d2c)
            return ec, sc, nm, rows

        tasks = [asyncio.ensure_future(_fetch_hist(*item)) for item in to_fetch]
        completed = 0
        for coro in asyncio.as_completed(tasks):
            if cancel_check and cancel_check():
                for t in tasks:
                    t.cancel()
                if on_progress:
                    on_progress(completed, progress_total,
                                f"[停止] ユーザーによる停止（{completed}/{fetch_total}件処理済み）")
                db.commit()
                yield None   # cancellation sentinel → driver returns early
                return
            edinet_code, sec_code, name, rows = await coro
            completed += 1
            if on_progress:
                prog = completed if backfill else skipped_total + completed
                on_progress(prog, progress_total,
                            f"[{prog}/{progress_total}] {name}({sec_code}) {len(rows) if rows else 0}件")
            yield [
                {"edinet_code": edinet_code, "trade_date": r["trade_date"],
                 "close": r.get("close"), "volume": r.get("volume")}
                for r in rows
            ] if rows else []

    async with httpx.AsyncClient(timeout=60) as session:
        cancelled, inserted_total = await _price_collection_driver(db, _stooq_batch_gen(session))

    if cancelled:
        return {"cancelled": True, "inserted": inserted_total, "skipped": skipped_total}
    if on_progress:
        on_progress(progress_total, progress_total,
                    f"[完了] {total}社処理（スキップ:{skipped_total}社）、{inserted_total}件追加")
    return {"cancelled": False, "inserted": inserted_total, "skipped": skipped_total, "companies": total}


async def _jquants_fetch_date(session: httpx.AsyncClient, api_key: str, date_str: str) -> list:
    """date_str (YYYY-MM-DD) の全銘柄 OHLCV を返す（ページネーション対応）。
    V2 API: x-api-key ヘッダーで認証。レスポンスキーは "data"、フィールドは O/H/L/C/Vo。
    400 = 非営業日またはサブスクリプション範囲外 → 空リストで正常終了。
    403 = カバレッジ外またはキー無効 → JQuantsCoverageError を送出（Issue #412）。
    429 = レート制限 → 90秒待って1回だけ再試行。それでも429 なら skip。
    """
    headers = {"x-api-key": api_key}
    rows = []
    pagination_key = None
    while True:
        params: dict = {"date": date_str}
        if pagination_key:
            params["pagination_key"] = pagination_key
        r = await session.get(JQUANTS_ENDPOINT, headers=headers, params=params, timeout=30)
        if r.status_code == 429:
            # 指数バックオフは429リトライがクォータを浪費するため使わない。
            # 90秒待って1回だけ再試行（60s ウィンドウが確実にリセットされる余裕）。
            log.warning(f"J-Quants 429: {date_str} → 90秒後に再試行")
            await asyncio.sleep(90)
            r = await session.get(JQUANTS_ENDPOINT, headers=headers, params=params, timeout=30)
            if r.status_code == 429:
                log.error(f"J-Quants 429: {date_str} → 再試行も429、スキップ")
                return []
        if r.status_code == 403:
            # 契約失効はカバレッジ境界と同じ 403 で返るが、ボディの文言で区別できる（#461）。
            body = (r.text or "").lower()
            raise JQuantsCoverageError(
                date_str, no_subscription=JQUANTS_NO_SUBSCRIPTION_MARK in body)
        if r.status_code in (400, 404):
            break  # 非営業日またはサブスクリプション範囲外
        r.raise_for_status()
        data = r.json()
        rows.extend(data.get("data", []))
        pagination_key = data.get("pagination_key")
        if not pagination_key:
            break
        await asyncio.sleep(JQUANTS_RATE_SLEEP)  # ページ間もレート制限を考慮
    return rows


async def _fetch_jquants_listed_info(session: httpx.AsyncClient, api_key: str) -> dict:
    """J-Quants /markets/listed/info から全上場銘柄の情報を取得する。

    戻り値: {"issued_shares": {sec_code(4桁): issued_shares(float)}, "active_codes": {sec_code(4桁), ...}}
    active_codes はレスポンスに現れた全銘柄（発行済株式数の有無を問わない）＝現在の上場銘柄集合
    （is_active 同期・Issue #315 に使う）。取得失敗時は両方空。
    """
    headers = {"x-api-key": api_key}
    empty = {"issued_shares": {}, "active_codes": set()}
    try:
        r = await session.get(JQUANTS_LISTED_INFO_ENDPOINT, headers=headers, timeout=30)
        if r.status_code == 429:
            await asyncio.sleep(90)
            r = await session.get(JQUANTS_LISTED_INFO_ENDPOINT, headers=headers, timeout=30)
        if not r.is_success:
            log.warning(f"J-Quants listed/info 取得失敗 status={r.status_code}")
            return empty
        issued_shares: dict = {}
        active_codes: set = set()
        for item in r.json().get("info", []):
            code = str(item.get("Code", ""))
            if not code:
                continue
            sec_code = code[:4]
            active_codes.add(sec_code)
            shares = item.get("IssuedShares")
            if shares is not None:
                issued_shares[sec_code] = float(shares)
        return {"issued_shares": issued_shares, "active_codes": active_codes}
    except Exception as e:
        log.warning(f"J-Quants listed/info 例外: {e}")
        return empty


def _update_issued_shares(db, sec_to_edinet: dict, issued_shares_map: dict) -> int:
    """companies.issued_shares を J-Quants 値で更新し、
    financial_records.issued_shares が NULL の最新レコードにも補完する。
    戻り値: 更新した companies 行数。
    """
    updated = 0
    for sec_code, shares in issued_shares_map.items():
        edinet_code = sec_to_edinet.get(sec_code)
        if not edinet_code or shares <= 0:
            continue
        rows = db.query(Company).filter(Company.edinet_code == edinet_code).all()
        for co in rows:
            co.issued_shares = shares
            updated += 1

    if updated:
        db.flush()
        # 最新の financial_record で issued_shares が NULL のものを J-Quants 値で補完
        from sqlalchemy import text as _text
        db.execute(_text("""
            UPDATE financial_records fr
            SET issued_shares = c.issued_shares
            FROM companies c
            WHERE fr.edinet_code = c.edinet_code
              AND c.issued_shares IS NOT NULL
              AND fr.issued_shares IS NULL
              AND fr.year = (
                  SELECT MAX(fr2.year)
                  FROM financial_records fr2
                  WHERE fr2.edinet_code = fr.edinet_code
              )
        """))
        db.commit()
        log.info(f"J-Quants 発行済株式数: {updated}社を companies に更新")
    return updated


async def collect_stock_price_history_jquants(
    db,
    days_back: int = 14,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    on_progress: Optional[Callable] = None,
    cancel_check: Optional[Callable] = None,
) -> dict:
    """J-Quants API から日次 OHLCV を日付単位で取得し ON CONFLICT UPDATE で保存する。
    J-Quants は JPX 公式データのため、stooq 由来レコードより優先して上書きする。
    1回のリクエストで全銘柄のデータが取得できるため stooq より大幅に高速。
    date_from/date_to を指定した場合はその範囲を使用し、省略時は days_back から計算する。
    プロバイダー固有ロジック（J-Quants 日付単位フェッチ）を _jquants_batch_gen に分離し、
    _price_collection_driver の共通フレームで DB 保存・trim を一元管理する。
    株式分割・併合を遡及反映した調整後値（Adj* フィールド）を使用する（Issue #314）。
    stooq/Yahoo（バックフィル経路）も調整済み系列のため、ソース間の整合性が取れる。
    """
    api_key = os.environ.get("JQUANTS_API_KEY", "")
    if not api_key:
        raise ValueError("環境変数 JQUANTS_API_KEY が未設定です")

    today  = date.today()
    _from  = date_from if date_from is not None else (today - timedelta(days=days_back))
    _to    = date_to   if date_to   is not None else today
    span   = (_to - _from).days + 1
    # 土日は J-Quants も空を返すのでスキップして API コール数を削減
    dates = [
        (_from + timedelta(days=i)).strftime("%Y-%m-%d")
        for i in range(span)
        if (_from + timedelta(days=i)).weekday() < 5
        and (_from + timedelta(days=i)) <= _to
    ]

    # sec_code (4桁) → edinet_code のルックアップ（全社一括で1回のみ）
    sec_to_edinet: dict = {
        row.sec_code: row.edinet_code
        for row in db.query(Company.sec_code, Company.edinet_code)
        .filter(Company.sec_code.isnot(None))
        .all()
    }

    total = len(dates)
    # forbidden: 403 で読み飛ばした日数（Issue #412）
    # no_subscription: 403 のボディが契約失効を明示した日数（#461）
    # aborted_days: 連続 403 の早期打ち切りで**叩かずに飛ばした**日数（#461）
    fetch_stats = {"forbidden": 0, "no_subscription": 0, "aborted_days": 0}

    async def _jquants_batch_gen(session):
        completed = 0
        last_req_time: float = 0.0
        consecutive_forbidden = 0
        for date_str in dates:
            # 連続 403 が続いたら残りを叩かない（#461）。失効・権限喪失は全日 403 になり、
            # 1日ごとに JQUANTS_RATE_SLEEP=20秒 を待つため窓の長さぶん丸ごと捨てることになる。
            if consecutive_forbidden >= JQUANTS_MAX_CONSECUTIVE_FORBIDDEN:
                fetch_stats["aborted_days"] = total - completed
                reason = ("契約失効（No active subscription）"
                          if fetch_stats["no_subscription"] else "カバー範囲外が連続")
                log.warning(
                    f"J-Quants: {consecutive_forbidden}日連続で403のため残り "
                    f"{fetch_stats['aborted_days']}日の取得を打ち切る（{reason}）"
                )
                if on_progress:
                    on_progress(completed, total,
                                f"[打ち切り] 403が{consecutive_forbidden}日連続（{reason}）"
                                f"・残り{fetch_stats['aborted_days']}日をスキップ")
                return
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(completed, total, f"[停止] ユーザーによる停止（{completed}/{total}日処理済み）")
                db.commit()
                yield None   # cancellation sentinel → driver returns early
                return

            # リクエスト開始間隔を最低 JQUANTS_RATE_SLEEP 秒に保つ（高速な祝日レスポンス後も適用）
            if completed > 0:
                elapsed = asyncio.get_event_loop().time() - last_req_time
                wait = JQUANTS_RATE_SLEEP - elapsed
                if wait > 0:
                    await asyncio.sleep(wait)

            last_req_time = asyncio.get_event_loop().time()
            try:
                quote_rows = await _jquants_fetch_date(session, api_key, date_str)
            except JQuantsCoverageError as e:
                # 無料プランのカバレッジ境界（例: days_back=730 の下限日）は 403 を返す。
                # 例外を伝播させると収集全体が落ちるため欠測扱いで継続する（Issue #412）。
                fetch_stats["forbidden"] += 1
                consecutive_forbidden += 1
                completed += 1
                if e.no_subscription:
                    # 契約失効はカバレッジ境界と別物＝**平常運転と読み違えない文言で出す**（#461）
                    fetch_stats["no_subscription"] += 1
                    log.error(
                        f"J-Quants 403: {date_str} — 契約が有効でない"
                        f"（No active subscription）。カバー範囲外ではなく購読状態の問題"
                    )
                else:
                    log.warning(f"J-Quants 403: {date_str} をカバー範囲外としてスキップ")
                if on_progress:
                    on_progress(completed, total,
                                f"[{completed}/{total}] {date_str} スキップ（403 "
                                f"{'契約失効' if e.no_subscription else 'カバー範囲外'}）")
                yield []
                continue
            consecutive_forbidden = 0
            completed += 1

            if not quote_rows:
                if on_progress:
                    on_progress(completed, total, f"[{completed}/{total}] {date_str} スキップ（非営業日）")
                yield []
                continue

            records = []
            for q in quote_rows:
                code        = str(q.get("Code", ""))
                sec_code    = code[:4]   # J-Quants は "13010"（5桁）→ 先頭4桁が証券コード
                edinet_code = sec_to_edinet.get(sec_code)
                if not edinet_code:
                    continue
                # V2: AdjC 等は株式分割・併合を遡及反映した調整後値（Issue #314）。
                # 未調整の C/O/H/L/Vo を使うと分割日を境に系列が段差になりリターン計算が破綻する。
                close_val = q.get("AdjC")
                if close_val is None:
                    continue   # close は nullable=False のためスキップ
                try:
                    records.append({
                        "edinet_code": edinet_code,
                        "sec_code":    sec_code,
                        "trade_date":  q["Date"],
                        "open":        float(q["AdjO"])  if q.get("AdjO")  is not None else None,
                        "high":        float(q["AdjH"])  if q.get("AdjH")  is not None else None,
                        "low":         float(q["AdjL"])  if q.get("AdjL")  is not None else None,
                        "close":       float(close_val),
                        "volume":      float(q["AdjVo"]) if q.get("AdjVo") is not None else None,
                    })
                except (KeyError, ValueError, TypeError):
                    continue

            # 同一 edinet_code に複数の J-Quants コードが対応する場合（優先株等）に
            # ON CONFLICT DO UPDATE の CardinalityViolation を防ぐため重複排除
            seen: set = set()
            deduped = []
            for rec in records:
                if rec["edinet_code"] not in seen:
                    seen.add(rec["edinet_code"])
                    deduped.append(rec)

            if on_progress:
                on_progress(completed, total, f"[{completed}/{total}] {date_str} {len(deduped)}件")
            yield deduped

    async with httpx.AsyncClient(timeout=60) as session:
        cancelled, upserted_total = await _price_collection_driver(db, _jquants_batch_gen(session))
        # 価格収集と同じセッションで上場銘柄情報（発行済株式数・現在の上場銘柄集合）も取得（API キー共用）
        listed_info = await _fetch_jquants_listed_info(session, api_key)

    # 403 の切り分け（Issue #412 → #425 で再設計 → #461 で失効を分離）:
    # #412 は「全日程 403 かつ listed/info も失敗ならキー失効とみなし中断」していた。#425 は
    # これを「`/markets/listed/info` は無料プランで常に 403」と解釈して撤去したが、**2026-08-07
    # の実測でその解釈は誤りと判明**——同エンドポイントは
    # `{"message": "The requested endpoint does not exist..."}` を返しており、**プランの制約では
    # なく v2 に存在しない URL** だった（正しい URL の確定は契約復旧後・#461）。いずれにせよ
    # active_codes は常に空で条件は `forbidden == total` に退化するため、#425 の判断
    # （J-Quants の失敗で例外を投げない）はそのまま維持する。J-Quants（収集元A）の失敗が、
    # 同じキーに依存しない Yahoo ギャップ補完（収集元B・株価鮮度の実質唯一の担い手）まで
    # 巻き添えでブロックするのは誤り。例外は投げず結果で返し、呼び出し側が継続可否を判断する。
    #
    # **失効は「カバー範囲外」と別文言で出す**（#461）。両者を同じ警告に潰していたため、
    # 契約が切れた 2026-08 も「エンバーゴ内なら正常」と読めるログが毎晩流れ続けていた。
    attempted = fetch_stats["forbidden"] + fetch_stats["aborted_days"]
    all_forbidden = bool(total) and attempted == total
    if fetch_stats["forbidden"]:
        log.warning(
            f"J-Quants 403: {fetch_stats['forbidden']}/{total}日をスキップ"
            + (f"（うち契約失効 {fetch_stats['no_subscription']}日）"
               if fetch_stats["no_subscription"] else "（カバー範囲外）")
            + (f"・連続403で残り {fetch_stats['aborted_days']}日を打ち切り"
               if fetch_stats["aborted_days"] else "")
        )
        if fetch_stats["no_subscription"]:
            log.error(
                "J-Quants の契約が有効ではありません（No active subscription）。"
                "カバー範囲外の 403 ではなく購読状態の問題で、**日付を変えても取得できません**。"
                "株価鮮度は Yahoo ギャップ補完が維持しますが、issued_shares / 上場状態同期 /"
                "会社予想開示は止まります（Issue #461）"
            )
        elif all_forbidden and not cancelled:
            log.error(
                f"J-Quants が試行した全 {attempted} 日で 403 を返しました。"
                "無料プランの84日エンバーゴ内の窓を叩いた場合は正常ですが、"
                "エンバーゴ外の日付を含むならキー失効/権限喪失の可能性があります（Issue #425）"
            )

    if listed_info["issued_shares"]:
        _update_issued_shares(db, sec_to_edinet, listed_info["issued_shares"])

    # 現在の上場銘柄集合と companies.is_active を同期（廃止銘柄検知・Issue #315）。
    # 取得失敗時（active_codes 空）は同期をスキップし、既存の is_active を誤って
    # 全件 delisted 化しないようにする。
    if listed_info["active_codes"]:
        sync_result = sync_active_status(db, listed_info["active_codes"])
        if sync_result["delisted"] or sync_result["reactivated"]:
            log.info(
                f"上場状態同期: 新規delisted={sync_result['delisted']}件, "
                f"復帰={sync_result['reactivated']}件"
            )

    if cancelled:
        return {"cancelled": True, "upserted": upserted_total,
                "forbidden": fetch_stats["forbidden"], "all_forbidden": all_forbidden,
                "no_subscription": fetch_stats["no_subscription"],
                "aborted_days": fetch_stats["aborted_days"]}
    if on_progress:
        on_progress(total, total, f"[完了] {total}日処理・{upserted_total}件追加/更新")
    return {"cancelled": False, "upserted": upserted_total, "days": total,
            "forbidden": fetch_stats["forbidden"], "all_forbidden": all_forbidden,
            "no_subscription": fetch_stats["no_subscription"],
            "aborted_days": fetch_stats["aborted_days"]}


def _update_market_data_latest(db) -> int:
    """point_in_time=False: 各社の最新レコードのみ、最新株価（daily優先）で更新する。"""

    subq = (
        db.query(
            StockPriceDaily.edinet_code,
            sqla_func.max(StockPriceDaily.trade_date).label("max_date"),
        )
        .group_by(StockPriceDaily.edinet_code)
        .subquery()
    )
    latest_price_rows = (
        db.query(StockPriceDaily.edinet_code, StockPriceDaily.close)
        .join(
            subq,
            (StockPriceDaily.edinet_code == subq.c.edinet_code)
            & (StockPriceDaily.trade_date == subq.c.max_date),
        )
        .all()
    )

    valid_ecs = [ec for ec, price in latest_price_rows if price and price > 0]
    latest_fin_by_ec = _fetch_latest_fin_by_ec(db, valid_ecs)

    updated = 0
    for edinet_code, price in latest_price_rows:
        if price is None or price <= 0:
            continue
        latest = latest_fin_by_ec.get(edinet_code)
        if not latest:
            continue
        _apply_price_to_record(latest, price)
        updated += 1
        if updated % PRICE_COMMIT_BATCH == 0:
            db.commit()

    db.commit()
    log.info(f"update_market_data_from_history: {updated}社を更新")
    return updated


def _update_market_data_point_in_time(db) -> int:
    """point_in_time=True: 全財務レコードを period_end 近傍の週次株価で更新し、
    最新レコードは現在株価で上書きする。"""

    # ── point_in_time=True: 全レコードを period_end 近傍の株価で更新 ─────────
    # financial_records を先にロードし、対象会社×日付範囲でフィルタした
    # weekly 行だけを取得する（全件メモリ展開を回避）。
    all_records = db.query(FinancialRecord).all()

    # 最新レコード（year最大）を社別にインデックス（最後の上書きステップで使用）。
    # 対象は **annual のみ**（#421）。H1 を混ぜると同一 year で先着順に決まってしまい、
    # 現在株価の上書きが H1 行へ吸われて annual の per/pbr が凍結する
    # （`financial_metrics` VIEW は period_type='annual' しか見ない）。
    # なお近傍探索そのものは H1 行にも適用する（下の dated_records は絞らない）。
    latest_by_ec: dict = {}
    for rec in all_records:
        if rec.period_type != "annual":
            continue
        ec = rec.edinet_code
        if ec not in latest_by_ec or (rec.year or 0) > (latest_by_ec[ec].year or 0):
            latest_by_ec[ec] = rec

    # period_end を持つレコードのみが近傍探索の対象
    dated_records = [r for r in all_records if r.period_end]
    if not dated_records:
        log.info("update_market_data_from_history(point_in_time): period_end ありレコードが空のためスキップ")
        # 最新レコードへの現在株価上書きだけ実施（期間探索なし）
        for ec, info in latest_prices(db, list(latest_by_ec.keys())).items():
            latest_price = info.get("price")
            if not latest_price or latest_price <= 0:
                continue
            _apply_price_to_record(latest_by_ec[ec], latest_price)
        db.commit()
        return 0

    # 対象会社・日付範囲を算出して weekly を必要範囲だけロード
    # ec_subq は Query をそのまま渡す（SQLAlchemy が SELECT サブクエリへ変換）
    period_ends = [r.period_end for r in dated_records]
    date_lo = (min(period_ends) - timedelta(days=MAX_GAP_DAYS)).isoformat()
    date_hi = (max(period_ends) + timedelta(days=MAX_GAP_DAYS)).isoformat()
    ec_q = (
        db.query(FinancialRecord.edinet_code)
        .filter(FinancialRecord.period_end.isnot(None))
        .distinct()
    )
    weekly_rows = (
        db.query(
            StockPriceWeekly.edinet_code,
            StockPriceWeekly.trade_date,
            StockPriceWeekly.close_last,
        )
        .filter(
            StockPriceWeekly.edinet_code.in_(ec_q),
            StockPriceWeekly.trade_date >= date_lo,
            StockPriceWeekly.trade_date <= date_hi,
            StockPriceWeekly.close_last > 0,
        )
        .all()
    )

    if not weekly_rows:
        log.info("update_market_data_from_history(point_in_time): stock_price_weekly が空のためスキップ")
        return 0

    # {edinet_code: sorted list of (trade_date_str, close)}
    history: dict = defaultdict(list)
    for ec, td, cl in weekly_rows:
        history[ec].append((td, cl))
    for ec in history:
        history[ec].sort()  # trade_date の昇順

    updated = 0
    for rec in dated_records:
        prices = history.get(rec.edinet_code)
        if not prices:
            continue

        try:
            target = rec.period_end
        except (ValueError, TypeError):
            continue

        dates = [p[0] for p in prices]
        price_dict = dict(prices)
        best_price = _nearest_price(dates, price_dict, target.isoformat(), MAX_GAP_DAYS)

        if best_price is None:
            continue

        _apply_price_to_record(rec, best_price)
        updated += 1
        if updated % PRICE_COMMIT_BATCH == 0:
            db.commit()

    # 最新レコードは現在株価で上書き（スクリーニング用）。daily（直近窓）を優先し
    # 無ければ weekly にフォールバックする latest_prices で最新終値を引く。
    for ec, info in latest_prices(db, list(latest_by_ec.keys())).items():
        latest_price = info.get("price")
        if not latest_price or latest_price <= 0:
            continue
        _apply_price_to_record(latest_by_ec[ec], latest_price)

    db.commit()
    log.info(f"update_market_data_from_history(point_in_time): {updated}レコードを更新")
    return updated


def update_market_data_from_history(db, point_in_time: bool = False) -> int:
    """stock_price_history の終値を financial_records.stock_price に反映する。
    stooq が GitHub Actions IP でブロックされる問題を回避するため、
    J-Quants 由来の stock_price_history を使ってバリュエーション指標を計算する。

    point_in_time=False（デフォルト・日次差分向け）:
        各社の最新レコードのみ、最新株価で更新する。高速。
    point_in_time=True（全件収集 finalize 向け）:
        全財務レコードを period_end 最近傍の株価で更新する。
        J-Quants カバレッジ外（データなし）のレコードはスキップし既存値を保持する。
        最新レコードは常に最新株価で上書きする。

    戻り値: 更新した財務レコード数
    """
    if not point_in_time:
        return _update_market_data_latest(db)
    return _update_market_data_point_in_time(db)


async def backfill_historical_stock_prices_yahoo(
    db,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
) -> int:
    """J-Quants カバー範囲外（JQUANTS_BACKFILL_DAYS 日より前）の financial_records で
    stock_price が NULL のレコードに対し、Yahoo Finance から period_end 近傍の
    株価を取得して financial_records.stock_price を直接更新する。

    stock_price_history には書き込まない（Supabase 500MB ストレージ節約のため）。
    J-Quants 由来の既存 stock_price は上書きしない（NULL のみ補完）。
    GitHub Actions（Azure IP）から動作する。
    """
    cutoff = date.today() - timedelta(days=JQUANTS_BACKFILL_DAYS)

    # 対象: stock_price が NULL かつ period_end が J-Quants カバー外
    target_records = (
        db.query(FinancialRecord)
        .filter(
            FinancialRecord.stock_price.is_(None),
            FinancialRecord.period_end.isnot(None),
            FinancialRecord.period_end < cutoff,
        )
        .all()
    )
    if not target_records:
        log.info("backfill_historical_stock_prices_yahoo: 対象レコードなし")
        return 0

    # edinet_code → sec_code マッピング
    sec_map = {
        c.edinet_code: c.sec_code
        for c in db.query(Company.edinet_code, Company.sec_code)
        .filter(Company.sec_code.isnot(None))
        .all()
    }

    # 企業ごとにグループ化（1企業=1 Yahoo リクエストで複数 period_end をカバー）
    by_company: dict = defaultdict(list)
    for rec in target_records:
        sec_code = sec_map.get(rec.edinet_code)
        if sec_code and sec_code.strip():
            by_company[sec_code].append(rec)

    total   = len(by_company)
    updated = 0

    async with httpx.AsyncClient(timeout=60) as session:
        for i, (sec_code, recs) in enumerate(sorted(by_company.items()), 1):
            if cancel_check and cancel_check():
                db.commit()
                if on_progress:
                    on_progress(i - 1, total, f"[Yahoo backfill] 停止（{updated}件更新済み）")
                return updated

            # この企業の全 period_end をカバーする日付範囲（±MAX_GAP_DAYS の余裕を持たせる）
            period_ends = sorted(r.period_end for r in recs)
            d_from = (period_ends[0]  - timedelta(days=MAX_GAP_DAYS)).strftime("%Y%m%d")
            d_to   = (period_ends[-1] + timedelta(days=MAX_GAP_DAYS)).strftime("%Y%m%d")

            # Yahoo Finance ティッカー（東証: {sec_code}.T）
            ticker = f"{sec_code}.T"
            rows = await fetch_yahoo_history(session, ticker, d_from, d_to)

            if rows:
                # {trade_date_str: close} の辞書
                price_dict = {r["trade_date"]: r["close"] for r in rows if r["close"]}
                price_dates = sorted(price_dict.keys())

                for rec in recs:
                    target_str = rec.period_end.isoformat() if rec.period_end else ""
                    best_price = _nearest_price(price_dates, price_dict, target_str, MAX_GAP_DAYS)
                    if best_price and best_price > 0:
                        _apply_price_to_record(rec, best_price)
                        updated += 1

                if updated % PRICE_COMMIT_BATCH == 0:
                    db.commit()

            if on_progress and (i % YAHOO_BACKFILL_PROGRESS_BATCH == 0 or i == total):
                on_progress(i, total, f"[Yahoo backfill {i}/{total}] {sec_code}  累計{updated}件更新")

            # レート制限対策（fetch_yahoo_history 内の処理時間は含まれるため短めのスリープ）
            if YAHOO_STOCK_RATE_SLEEP > 0:
                await asyncio.sleep(YAHOO_STOCK_RATE_SLEEP)

    db.commit()
    log.info(f"backfill_historical_stock_prices_yahoo: {updated}件の financial_records を更新")
    return updated


async def fill_recent_stock_price_gap_yahoo(
    db,
    gap_days: int = 7,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
) -> dict:
    """各社の株価が gap_days 日以上古い場合に、Yahoo Finance から不足期間を
    **銘柄ごとに**補完して株価テーブルへ追記する。
    差分収集（incremental）後のフォールバックとして使用。
    J-Quants データが存在する行は上書きしない（ON CONFLICT DO NOTHING）。

    起点は銘柄別の最終 trade_date（Issue #415）。全社横断の最大日を1つ選んで全社へ
    適用すると、収集が数日止まった後に一部銘柄だけ先行して復旧した場合、遅れている
    大多数の銘柄の欠測期間が永久に埋まらない（2026-07 に実際に発生。2銘柄が 07-31 /
    3,677銘柄が 07-13 の状態で d_from=08-01 となり 14営業日分が穴のまま残った）。
    週次バーの欠落は例外を出さず、build_snapshots の 52週先ラベル（インデックス参照）
    や px_* のローリング窓を静かにずらすため検知が難しい。
    per-company 起点にしても Yahoo は元々1社1リクエストのため**リクエスト数は増えない**
    （遅延銘柄の日付レンジが広がるだけ）。同じ per-company 判定は
    backfill_weekly_history_yahoo が既に採っている。

    プロバイダー固有ロジック（Yahoo 逐次フェッチ）を _yahoo_batch_gen に分離し、
    _price_collection_driver の共通フレームで DB 保存・trim を一元管理する。
    """
    today = date.today()

    # 銘柄別の最終 trade_date。daily は保持窓（DAILY_WINDOW_DAYS）で trim されるため、
    # 長期停止した社は weekly 側にしか残らない。両者の新しい方を起点にする。
    latest_daily = dict(
        db.query(StockPriceDaily.edinet_code, sqla_func.max(StockPriceDaily.trade_date))
        .group_by(StockPriceDaily.edinet_code).all()
    )
    latest_weekly = dict(
        db.query(StockPriceWeekly.edinet_code, sqla_func.max(StockPriceWeekly.trade_date))
        .group_by(StockPriceWeekly.edinet_code).all()
    )
    if not latest_daily and not latest_weekly:
        log.info("fill_recent_stock_price_gap_yahoo: 株価データが空のためスキップ")
        return {"skipped": True, "reason": "empty"}

    # 起点の下限。これより過去への遡及は backfill_weekly_history_yahoo の管轄
    # （daily 保持窓を超える取得を毎日走らせない＝暴走ガード）。
    floor_d = today - timedelta(days=DAILY_WINDOW_DAYS)

    companies = [
        (row.sec_code, row.edinet_code)
        for row in db.query(Company.sec_code, Company.edinet_code)
        .filter(Company.sec_code.isnot(None))
        .all()
    ]

    to_fetch = []
    for sec_code, edinet_code in companies:
        _d, _w = latest_daily.get(edinet_code), latest_weekly.get(edinet_code)
        last = max(x for x in (_d, _w) if x) if (_d or _w) else None
        if last is None:
            start = floor_d                       # 株価未収集の社は保持窓の先頭から
        else:
            last_d = date.fromisoformat(last[:10])
            if (today - last_d).days <= gap_days:
                continue                          # この社はギャップなし
            start = max(last_d + timedelta(days=1), floor_d)
        to_fetch.append((sec_code, edinet_code, start.strftime("%Y%m%d")))

    total = len(to_fetch)
    if total == 0:
        log.info(f"fill_recent_stock_price_gap_yahoo: 全 {len(companies)} 社ギャップなし")
        return {"skipped": True, "reason": "no_gap", "companies": 0}

    to_fetch.sort()          # 銘柄順に部分反映されるよう順序を決定的にする
    d_from_min = min(x[2] for x in to_fetch)
    d_to = today.strftime("%Y%m%d")
    log.info(f"fill_recent_stock_price_gap_yahoo: {total}/{len(companies)}社を補完"
             f"（最古起点 {d_from_min} 〜 {d_to}）")

    async def _yahoo_batch_gen(session):
        for i, (sec_code, edinet_code, d_from) in enumerate(to_fetch, 1):
            rows = await fetch_yahoo_history(session, f"{sec_code}.T", d_from, d_to)
            # ギャップ補完は各社の最終日より後の新規日付が対象のため衝突は稀。
            records = [
                {"edinet_code": edinet_code, "trade_date": r["trade_date"],
                 "close": r["close"], "volume": r.get("volume")}
                for r in rows if r["close"]
            ] if rows else []
            if on_progress and i % PROGRESS_REPORT_BATCH == 0:
                on_progress(i, total, f"[Yahoo gap-fill {i}/{total}]")
            yield records
            await asyncio.sleep(YAHOO_STOCK_RATE_SLEEP)

    async with httpx.AsyncClient(timeout=60) as session:
        _, upserted = await _price_collection_driver(db, _yahoo_batch_gen(session))

    log.info(f"fill_recent_stock_price_gap_yahoo: {upserted}件を株価テーブルへ集約保存")
    return {"skipped": False, "upserted": upserted, "companies": total,
            "from": d_from_min, "to": d_to}


async def backfill_weekly_history_yahoo(
    db,
    years_back: int = 5,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
) -> dict:
    """stock_price_weekly を過去方向へ years_back 年まで延伸する（Yahoo Finance / #198）。

    背景: use_momentum=ON（macro_risk_return §9.8）は 52週先リターン＋12ヶ月モメンタムを
    同時に要求するため、週次株価の被覆が短いと walk-forward CV が 0 フォルドになる。
    各社の週次の最古 trade_date が today-years_back より新しい（＝過去が不足）場合に、
    不足期間を Yahoo から取得し record_prices_batch 経由で daily→weekly 再集約して埋める。

    ストレージ安全性: 1社処理ごとに record_prices_batch(trim=True) を呼び、daily を保持窓
    （DAILY_WINDOW_DAYS）以外は都度 trim する。これにより 5年×全社の daily が同時展開して
    Supabase Free 500MB を超えるのを防ぐ。weekly は古い daily から再集約済みのため情報損失なし。
    既存 weekly 行は ON CONFLICT UPDATE で同値上書き（破壊なし）。
    J-Quants カバー外の過去も Yahoo で取得でき、GitHub Actions（Azure IP）から動作する。
    """
    today     = date.today()
    floor_d   = date(today.year - years_back, today.month, today.day)
    floor_str = floor_d.isoformat()
    d_from    = floor_d.strftime("%Y%m%d")

    # 企業ごとの週次最古 trade_date（weekly 未収集の社はキー無し）
    min_week = dict(
        db.query(StockPriceWeekly.edinet_code,
                 sqla_func.min(StockPriceWeekly.trade_date))
        .group_by(StockPriceWeekly.edinet_code)
        .all()
    )

    companies = (
        db.query(Company.sec_code, Company.edinet_code)
        .filter(Company.sec_code.isnot(None), Company.sec_code != "")
        .all()
    )

    # 取得対象: weekly 未収集、または最古日が floor より新しい（過去が不足する）社のみ
    to_fetch = []
    for sec_code, edinet_code in companies:
        oldest = min_week.get(edinet_code)
        if oldest is None:
            d_to = today
        elif oldest > floor_str:
            d_to = date.fromisoformat(oldest) - timedelta(days=1)
        else:
            continue  # 既に years_back 以上カバー済み
        to_fetch.append((sec_code, edinet_code, d_to.strftime("%Y%m%d")))

    total = len(to_fetch)
    if total == 0:
        log.info(f"backfill_weekly_history_yahoo: 全社 {years_back}年以上カバー済み（対象なし）")
        return {"skipped": True, "reason": "already_covered", "companies": 0}

    upserted = 0
    async with httpx.AsyncClient(timeout=60) as session:
        for i, (sec_code, edinet_code, d_to) in enumerate(sorted(to_fetch), 1):
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(i - 1, total, f"[週次backfill] 停止（{upserted}件保存済み）")
                return {"cancelled": True, "upserted": upserted, "companies": i - 1}

            rows = await fetch_yahoo_history(session, f"{sec_code}.T", d_from, d_to)
            records = [
                {"edinet_code": edinet_code, "trade_date": r["trade_date"],
                 "close": r["close"], "volume": r.get("volume")}
                for r in rows if r.get("close")
            ] if rows else []
            if records:
                try:
                    # 1社ごとに trim=True：daily を都度 trim して保持窓外の過去を残さない
                    upserted += record_prices_batch(db, records, trim=True)
                except Exception as e:
                    log.warning("週次backfill バッチ保存失敗 %s: %s", sec_code, e)

            if on_progress and (i % YAHOO_BACKFILL_PROGRESS_BATCH == 0 or i == total):
                on_progress(i, total, f"[週次backfill {i}/{total}] {sec_code} 累計{upserted}件")

            if YAHOO_STOCK_RATE_SLEEP > 0:
                await asyncio.sleep(YAHOO_STOCK_RATE_SLEEP)

    log.info(f"backfill_weekly_history_yahoo: {upserted}件の daily を保存し weekly を再集約（{total}社）")
    return {"skipped": False, "upserted": upserted, "companies": total, "floor": floor_str}


def _nearest_price(sorted_dates: list, price_dict: dict, target_str: str,
                   max_gap: int) -> Optional[float]:
    """昇順の日付文字列リスト `sorted_dates` から `target_str` に最も近い日付の
    価格（`price_dict[日付]`）を返す。最近傍の日付差が `max_gap` 日を超える場合は
    None。bisect で挿入位置を求め、その前後2候補のみを比較する。

    point-in-time マッチと Yahoo backfill の最近傍探索で共用する内部ヘルパー。
    """
    try:
        target = date.fromisoformat(target_str[:10])
    except (ValueError, TypeError):
        return None

    pos = bisect.bisect_left(sorted_dates, target_str)
    best_price = None
    best_gap = max_gap + 1

    for idx in (pos - 1, pos):
        if 0 <= idx < len(sorted_dates):
            td_str = sorted_dates[idx]
            try:
                td = date.fromisoformat(td_str[:10])
            except ValueError:
                continue
            gap = abs((td - target).days)
            if gap < best_gap:
                best_gap = gap
                best_price = price_dict[td_str]

    return best_price


def _apply_price_to_record(rec, price: float) -> None:
    """財務レコードに株価・バリュエーション指標を書き込む（内部ヘルパー）"""
    rec.stock_price = price
    if rec.pl_eps and rec.pl_eps > 0:
        rec.per = round(price / rec.pl_eps, 2)
    if rec.bs_bps and rec.bs_bps > 0:
        rec.pbr = round(price / rec.bs_bps, 2)
    _sh = (float(rec.issued_shares) if (rec.issued_shares and rec.issued_shares > 0)
           else ((rec.bs_total_equity / rec.bs_bps)
                 if (rec.bs_bps and rec.bs_bps > 0
                     and rec.bs_total_equity and rec.bs_total_equity > 0)
                 else None))
    if _sh:
        rec.market_cap = round(price * _sh / 1_000_000, 2)
    if rec.dps and rec.dps > 0 and price > 0:
        rec.div_yield = round(rec.dps / price * 100, 2)


def _fetch_latest_fin_by_ec(db, edinet_codes: list) -> dict:
    """各社の最新 **通期（annual）** FinancialRecord を1クエリで取得して
    {edinet_code: record} を返す。

    ROW_NUMBER() OVER (PARTITION BY edinet_code ORDER BY year DESC, period_end DESC)
    で最新行を確定するため、同一 year・複数 period_end が存在しても安全。
    N+1 クエリの代替として update_market_data_from_history から使用。

    `period_type == 'annual'` で絞るのは、株価・per/pbr/market_cap の書き込み先を
    `financial_metrics` VIEW（`WHERE fr.period_type = 'annual'`）と揃えるため（#421）。
    絞らないと同一 year に H1 が併存する社で period_end の新しい H1 行が「最新」に
    選ばれ、更新が H1 へ吸われて **annual 行の per/pbr が凍結する**（VIEW は H1 を
    見ないので画面上は「古い株価のまま」に見える）。H1 の period_end が annual を
    追い越すケースは #424 子1（H1 の定期収集）で増えるため、先に塞いでおく。
    """
    if not edinet_codes:
        return {}
    rn = sqla_func.row_number().over(
        partition_by=FinancialRecord.edinet_code,
        order_by=[FinancialRecord.year.desc(), FinancialRecord.period_end.desc()],
    ).label("rn")
    subq = (
        db.query(FinancialRecord.id, rn)
        .filter(FinancialRecord.edinet_code.in_(edinet_codes))
        .filter(FinancialRecord.period_type == "annual")
        .subquery()
    )
    return {
        r.edinet_code: r
        for r in db.query(FinancialRecord)
        .join(subq, (FinancialRecord.id == subq.c.id) & (subq.c.rn == 1))
        .all()
    }


# ── マクロデータ（為替・金利・指数・コモディティ）─────────────────────────

# ── 観測基準日（anchor）の群既定 ────────────────────────────────────────────
# `lag_days` は**この日に加算**されて trade_date になる。基準日が群ごとに違うため、
# **`lag_days` の数値を群をまたいで直接比較してはならない**（CONTEXT.md「公表ラグ補正」）。
# 四半期系列なら期首基準と期末基準で 90 日ぶんの下駄が付き、JP_REAL_GDP（FRED・135）と
# ESRI GDP（60）はどちらも実質「期末+45〜60日」を指す——この差が明文化されていなかった
# ことが #447 の旧診断（四半期系列に 80〜96 日の先読みがあるという誤り）を招いた。
#
#   period_start: 参照期間の**期首**日を基準にする（FRED / 日銀 / e-Stat / OECD）
#   period_end:   参照期間の**期末**＝翌期初日を基準にする（ESRI・_parse_esri_gdp_csv）
#   collection:   参照期間を持たず**収集日**そのものを使う（市場系・ニュース系・IMF 継続収集）
#
# 新しい収集チャネルを足すときはここへ 1 行加える（`tests/test_collect_macro.py::
# test_every_series_group_declares_anchor` が未宣言を落とす）。
SERIES_ANCHOR: dict[str, str] = {
    "MACRO_SERIES":       "collection",    # Yahoo/stooq の市場系（その日の終値）
    "FRED_SERIES":        "period_start",
    "BOJ_SERIES":         "period_start",
    "OECD_SERIES":        "period_start",
    "ESTAT_SERIES":       "period_start",
    "ESTAT_INDEX_SERIES": "period_start",
    "ESRI_SERIES":        "period_end",    # 参照四半期の翌四半期初日（下記 _parse_esri_gdp_csv）
    "IMF_SERIES":         "period_start",  # バックフィルは vintage 期首。継続収集は収集日
    "GDELT_SERIES":       "collection",
    "WIKIMEDIA_SERIES":   "collection",
    "MOF_SERIES":         "period_start",  # CSV の「基準日」＝観測日。lag_days で公表時点へ寄せる
}

# stooq ティッカー定義。category は 'fx' / 'rate' / 'equity' / 'commodity' / 'volatility'。
# 本番収集は GitHub Actions（Azure IP）上で Yahoo Finance を優先する（stooq は 403 ブロック）。
# VIX/DXY/US5Y/US30Y は #218 フェーズ1 で追加。Yahoo のみで取得するため stooq ticker は
# best-effort（空文字は stooq フォールバック時に「データ無し」で skip され安全）。これらが
# macro_data に実際に蓄積されたことを Actions で実証してから M-1 の特徴量（_MACRO_MAP）へ公開する。
MACRO_SERIES: list[dict] = [
    {"code": "USDJPY",    "name": "USD/JPY",      "category": "fx",         "ticker": "usdjpy",   "yf_ticker": "USDJPY=X"},
    {"code": "EURJPY",    "name": "EUR/JPY",      "category": "fx",         "ticker": "eurjpy",   "yf_ticker": "EURJPY=X"},
    {"code": "DXY",       "name": "ドル指数",     "category": "fx",         "ticker": "",         "yf_ticker": "DX-Y.NYB"},
    {"code": "US5Y",      "name": "米5年金利",    "category": "rate",       "ticker": "",         "yf_ticker": "^FVX"},
    {"code": "US10Y",     "name": "米10年金利",   "category": "rate",       "ticker": "10usy.b",  "yf_ticker": "^TNX"},
    {"code": "US30Y",     "name": "米30年金利",   "category": "rate",       "ticker": "",         "yf_ticker": "^TYX"},
    # 日10年金利（JP10Y）は Yahoo `^JGB` 廃止（404）・stooq `10jpy.b` も0件で、**定義だけが残って
    # 1行も蓄積されない**系列だった（毎回2リクエスト空振りしてから「データ無し」で continue）。
    # #442 で削除。既定モデルは月次 FRED の `JP10Y_FRED`（下記）を使うため影響は無い。日次ソースは
    # 財務省「国債金利情報」CSV で取れることを #456 で確認済みだが、CSV パース経路のため
    # Yahoo/stooq 前提の本リストには乗らない。収集の実装は #458。
    {"code": "NIKKEI225", "name": "日経225",      "category": "equity",     "ticker": "^nkx",     "yf_ticker": "^N225"},
    # TOPIX 指数 ^TPX は Yahoo で配信停止（200 OK だが 0 件）。TOPIX 連動 ETF 1306.T
    # （NEXT FUNDS TOPIX・最長履歴・高流動）を代理に使う＝yoy/logret/zscore は指数と同等に追従。
    {"code": "TOPIX",     "name": "TOPIX",        "category": "equity",     "ticker": "^tpx",     "yf_ticker": "1306.T"},
    {"code": "SP500",     "name": "S&P500",       "category": "equity",     "ticker": "^spx",     "yf_ticker": "^GSPC"},
    {"code": "VIX",       "name": "VIX恐怖指数",  "category": "volatility", "ticker": "",         "yf_ticker": "^VIX"},
    {"code": "WTI",       "name": "WTI原油",      "category": "commodity",  "ticker": "cl.f",     "yf_ticker": "CL=F"},
    {"code": "GOLD",      "name": "金",           "category": "commodity",  "ticker": "gc.f",     "yf_ticker": "GC=F"},
    # ── コモディティ・チャネル拡張（#358・ADR-0013）───────────────────────────
    # 日本株の業種別コモディティ感応度をカバー（銅=非鉄/電線/機械・天然ガス=電力ガス/化学・
    # 貴金属=商社/触媒/電子材料・穀物=食品/飼料）。Phase 0 疎通検証で全8系列が Yahoo v8 から
    # 6年 1506-1510 行取得可。stooq はコモディティ先物が全滅（ローカル IP でも 0 件）のため
    # ticker は空文字（Yahoo フォールバック時に安全に skip）。VIX/DXY 同様、macro_data への
    # 蓄積を Actions で実証してから M-1/M-2/M-3 の特徴量（_MACRO_MAP・_DLM_MACRO_MAP）へ
    # 公開する（2PR 構成・#218 の公開フロー準拠）。
    #
    # BCOM は指数 `^BCOM` が 2026-07-17 を最後に配信停止（#438）。以降も timestamp 行だけは
    # 返るが close が全 null になるため `fetch_yahoo_history` は 0 件扱いで continue し、
    # 例外もエラーも出さないまま `macro_bcom_yoy`（M-1/M-2/M-6 既定）と `dlm_bcom`（M-3 既定）
    # が固定値を返し続けていた。TOPIX `^TPX`→`1306.T`（#250）と同じく**連動商品を代理**にする:
    # `DJP`（iPath Bloomberg Commodity ETN）は 2016-07-25 以降 BCOM の全 2,509 営業日を
    # カバーし、追従性は週次 logret corr 0.9906 / yoy corr 0.9957（2026-08-06 実測。次点は
    # DBC 0.9027・BCI 0.836）。`^BCOMTR`/`^DJUBS` は Yahoo で死亡、`^SPGSCI` は生存だが
    # エネルギー偏重で WTI と重複＝ADR-0013 が却下済み。
    # **series_code は BCOM のまま**（特徴量名・既定セットは不変）。DJP はトータルリターン型
    # ＝旧 ER 指数と水準体系が違うため、切替時は全期間（--years 11）で再収集して旧行を
    # 上書きすること。期間を絞ると新旧 level が混在し yoy が1年間壊れる。
    {"code": "BCOM",      "name": "ブルームバーグ商品指数", "category": "commodity", "ticker": "", "yf_ticker": "DJP"},
    {"code": "COPPER",    "name": "銅先物",        "category": "commodity",  "ticker": "",         "yf_ticker": "HG=F"},
    {"code": "NATGAS",    "name": "天然ガス先物",  "category": "commodity",  "ticker": "",         "yf_ticker": "NG=F"},
    {"code": "SILVER",    "name": "銀先物",        "category": "commodity",  "ticker": "",         "yf_ticker": "SI=F"},
    {"code": "WHEAT",     "name": "小麦先物",      "category": "commodity",  "ticker": "",         "yf_ticker": "ZW=F"},
    {"code": "CORN",      "name": "トウモロコシ先物", "category": "commodity", "ticker": "",       "yf_ticker": "ZC=F"},
    {"code": "SOYBEAN",   "name": "大豆先物",      "category": "commodity",  "ticker": "",         "yf_ticker": "ZS=F"},
    {"code": "PLATINUM",  "name": "プラチナ先物",  "category": "commodity",  "ticker": "",         "yf_ticker": "PL=F"},
]

# ── FRED マクロ系列（クレジット・インフレ・JP金利・期間構造）──────────────────────────
# FRED_API_KEY が設定されている場合のみ収集。未設定時は collect_macro_data 内でスキップ。
# アカウント登録: https://fred.stlouisfed.org/docs/api/api_key.html （無料・要ユーザー登録）
# GitHub Actions シークレット名: FRED_API_KEY
# レート制限: 120 req/min → FRED_RATE_SLEEP=0.6s でバッファ込み
FRED_API_KEY  = os.getenv("FRED_API_KEY", "")
FRED_BASE_URL = "https://api.stlouisfed.org/fred/series/observations"
FRED_RATE_SLEEP = 0.6  # 120 req/min = 2/s → 0.6s でバッファ付き

FRED_SERIES: list[dict] = [
    {"code": "HY_OAS",       "name": "米HYスプレッド（OAS）",     "category": "credit",    "fred_id": "BAMLH0A0HYM2"},
    {"code": "IG_OAS",       "name": "米IGスプレッド（OAS）",     "category": "credit",    "fred_id": "BAMLC0A0CM"},
    # 非ICE代替の信用スプレッド（#381）。FRED は 2026-04 以降 ICE BofA 系列（HY_OAS/IG_OAS）を
    # ローリング3年窓に制限し 2023-06 以前を配信しないため、strict M-1 の学習窓が24ヶ月に律速される。
    # BAA10Y（Moody's Baa 社債利回り − 米10年金利）は Moody's 由来で truncate されず日次1986〜取得可能。
    # M-1/M-4 の既定信用ファクターをこちらへ移し、HY_OAS/IG_OAS は選択肢（直近3年窓）として残す。
    {"code": "BAA_SPREAD",   "name": "米Baa社債スプレッド（Baa−10Y）", "category": "credit", "fred_id": "BAA10Y"},
    {"code": "BREAKEVEN10Y", "name": "米10年BEI（インフレ期待）", "category": "inflation", "fred_id": "T10YIE"},
    # IRLTLT01JPM156N は OECD MEI 由来の**月次**系列（末尾 M156N）。freq を省くと
    # _GROUP_DEFAULT_FREQ["FRED_SERIES"]="daily"（許容14日）で判定され、平常運転の
    # 月次ラグ（実測 35〜64日）が毎晩 CRITICAL になる＝#444 の狼少年。
    # lag_days=64（#447 で 70 から是正・anchor=period_start）: **実配信ラグの実測上限**。
    # 2026-08-04 時点で6月分（obs=2026-06-01）は配信済み・7月分は未配信＝6月分は obs+64 には
    # 確実に存在した。この系列は #444 の再収集で全行の `created_at` が潰れており実測点が無い
    # （ADR-0028 規則6 の罠にかかった実例）ため、先読みゼロを保証できるのは上限側のこの値だけ。
    # 参考の推定値は 55（月中平均の確定31 + OECD MEI→FRED 取り込み20〈JP_UNEMP で実測〉+ 4）
    # だが、推定で下げると検証手段の無い先読みが残るため採らない。
    # **70 は上限 64 すら超えており trade_date が常に未来日だった**（max=2026-08-10＝6日先）。
    # 未来日の間は経過日数が負になり stale 判定が構造的に成立しない（ADR-0028 Consequences）。
    # stale_days=62: 理論最大 31日（観測周期31 + 実配信ラグ64 − lag_days 64）+ 観測周期31。
    {"code": "JP10Y_FRED",   "name": "日10年金利（FRED）",        "category": "rate",      "fred_id": "IRLTLT01JPM156N", "freq": "monthly", "lag_days": 64, "stale_days": 62},
    {"code": "T10Y2Y",       "name": "米10y−2yスプレッド",       "category": "rate",      "fred_id": "T10Y2Y"},
    # ── 政策不確実性（#404）─────────────────────────────────────────────────
    # Baker-Bloom-Davis の Economic Policy Uncertainty Index（新聞記事ベース）。日次版のため
    # lag_days / freq は不要（#379/#382 の低頻度変換窓・#381 の strict 律速をどちらも回避）。
    # 1985-01-01 開始・現役更新（2026-07-30 に fredgraph.csv で最終観測日を実測確認）。
    # 日本版 JPNEPUINDXM は FRED 側で 2016-04 凍結のため採らない（#253 の JP_IP と同型）。
    {"code": "US_EPU",        "name": "米 経済政策不確実性指数（日次）", "category": "uncertainty", "fred_id": "USEPUINDXD"},
    {"code": "US_EQUITY_EPU", "name": "米 株式市場関連 経済不確実性指数", "category": "uncertainty", "fred_id": "WLEMUINDXD"},
    # ── 日本 実体経済指標（#250・日本マクロのリバランス）─────────────────────────
    # FRED は観測値を「期の参照開始日」の日付で返す。実体経済指標は公表ラグが大きい
    # （GDP=期末から約1.5〜2か月、月次指標=約1か月）ため、lag_days 分だけ trade_date を
    # 後ろへシフトして「この日には知れた値」へ正規化する＝先読みバイアス（look-ahead）防止。
    # lag_days 未指定の既存5系列は 0=シフト無し（完全後方互換）。
    # 採用前に各 fred_id の最終更新日を確認すること（OECD 旧系列は凍結あり：CPALTT01JPM657N 等）。
    {"code": "JP_REAL_GDP",  "name": "日本 実質GDP",        "category": "real_economy", "fred_id": "JPNRGDPEXP",      "freq": "quarterly", "lag_days": 135},
    # lag_days=82（#447 で 60 から是正・anchor=period_start）: 労働力調査は調査月の翌月末に
    # 公表される（期首基準の最遅 62日）が、**FRED（OECD MEI 経由）への取り込みはさらに遅れる**。
    # 2026年5月分は総務省が 06-30 に公表し、FRED へ現れたのは 07-20（macro_data.created_at 実測
    # ＝期首から 80日）。62 + 20 = 82。総務省の公表日で線を引くと取り込み分がそのまま先読みに
    # なる——**先読みは収集元（FRED）の実配信ラグで判定する**（ADR-0028 規則4・CONTEXT.md
    # 「実配信ラグ」）。
    # stale_days=62: 理論最大 31日（観測周期31 + 実配信ラグ82 − lag_days 82）+ 観測周期31。
    {"code": "JP_UNEMP",     "name": "日本 失業率",         "category": "labor",        "fred_id": "LRUNTTTTJPM156S", "freq": "monthly",   "lag_days": 82, "stale_days": 62},
    # JP_IP (JPNPROINDMISMEI) は 2024-04-30 で凍結確認済み (#253)。e-Stat コネクタ実装まで除外。
    {"code": "JP_TRADE_BAL", "name": "日本 貿易収支",       "category": "trade",        "fred_id": "XTNTVA01JPQ664S", "freq": "quarterly", "lag_days": 135},
]
# FRED 低頻度系列の履歴確保（四半期 zscore は ≥20 点必要・[macro_snapshots]_macro_from_cache）。
# 市場系（years_back）より長く遡って観測点を担保する。
FRED_MIN_YEARS_BACK = 10

# ── 日銀 時系列統計 API（stat-search.boj.or.jp/api/v1）─────────────────────
# 認証不要・JSON。ADR-0006 §Decision-2。
# 注: ADR は api.boj.or.jp と記したが実エンドポイントは stat-search.boj.or.jp/api/v1
#   （2026-02 新 API 発表後も stat-search が正式エンドポイント）→ GOTCHAS.md に記載済み。
BOJ_BASE_URL  = "https://www.stat-search.boj.or.jp/api/v1"
BOJ_RATE_SLEEP = 0.5  # 同一 DB への連続リクエストに備えたバッファ

# freq="monthly" → SURVEY_DATES は YYYYMM（e.g. 202501）
# freq="quarterly" → SURVEY_DATES は YYYYQQ（01=Q1/4月, 02=Q2/7月, 03=Q3/10月, 04=Q4/翌1月）
#
# 公表日＝実配信日（日銀の時系列統計 API は公表と同時に反映される。2026年6月分の実測で
# M2 が 07-09・CGPI が 07-10 に `macro_data.created_at` へ現れ、下記カレンダーと一致した）:
#   マネーストック速報    翌月第7営業日（3月・9月分は金融機関からの入手が遅れ第9営業日）
#   マネタリーベース      翌月第2営業日
#   企業物価指数（CGPI）  翌月第8営業日（速報値）
#   短観                  調査月の初旬（4/7/10/12月）＝ lag_days=14 は保守側で余裕あり
# lag_days は anchor=period_start（SERIES_ANCHOR）＝参照月の**期首**へ加算される。月次3系列は
# 「月の最大日数 31 + 正月休みを挟む1月の第N営業日」を実配信ラグの理論最大として置き直した（#447）。
# stale_days は月次3系列とも 62日（ADR-0028 規則5）。lag_days を実配信ラグの理論最大以上に
# 取っているため健全時の理論最大遅延は観測周期 31 日を超えず、その 2 倍＝「公表を1回
# スキップしたら鳴る」線になる。
BOJ_SERIES: list[dict] = [
    {
        "code": "JP_M2",
        "name": "日本 M2（マネーストック）",
        "category": "money",
        "db": "MD02",
        "boj_code": "MAM1NAM2M2MO",
        "freq": "monthly",
        # 理論最大 45日＝31（月の最大日数）+ 14（正月休みを挟む1月の第7営業日）。マージン込み 47。
        # 実測 2026年6月分は 07-09 公表＝期首から 38日。旧値 21 は **17日の先読み**だった（#447）。
        "lag_days": 47,
        "stale_days": 62,
    },
    {
        "code": "JP_TANKAN_MFG_LARGE",
        "name": "短観 製造業大企業 業況DI",
        "category": "survey",
        "db": "CO",
        "boj_code": "TK99F1000601GCQ01000",
        "freq": "quarterly",
        "lag_days": 14,
    },
    {
        "code": "JP_TANKAN_NONMFG_LARGE",
        "name": "短観 非製造業大企業 業況DI",
        "category": "survey",
        "db": "CO",
        "boj_code": "TK99F2000601GCQ01000",
        "freq": "quarterly",
        "lag_days": 14,
    },
    {
        "code": "JP_TANKAN_MFG_SMALL",
        "name": "短観 製造業中小企業 業況DI",
        "category": "survey",
        "db": "CO",
        "boj_code": "TK99F1000601GCQ03000",
        "freq": "quarterly",
        "lag_days": 14,
    },
    {
        "code": "JP_TANKAN_NONMFG_SMALL",
        "name": "短観 非製造業中小企業 業況DI",
        "category": "survey",
        "db": "CO",
        "boj_code": "TK99F2000601GCQ03000",
        "freq": "quarterly",
        "lag_days": 14,
    },
    {
        "code": "JP_CGPI",
        "name": "日本 企業物価指数（国内・総平均）",
        "category": "price",
        "db": "PR01",
        "boj_code": "PRCG20_2200000000",
        "freq": "monthly",
        # 理論最大 46日＝31 + 15（正月休みを挟む1月の第8営業日）。マージン込み 47。
        # 実測 2026年6月分は 07-10 公表＝期首から 39日。旧値 30 は **9日の先読み**だった（#447）。
        "lag_days": 47,
        "stale_days": 62,
    },
    {
        "code": "JP_MONETARY_BASE",
        "name": "日本 マネタリーベース平均残高",
        "category": "money",
        "db": "MD01",
        "boj_code": "MABS1AN11",
        "freq": "monthly",
        # 理論最大 37日＝31 + 6（正月休みを挟む1月の第2営業日）。マージン込み 38。
        # 2026年6月分の公表は 07-02（翌月第2営業日）＝期首から 31日。旧値 14 は **17日の先読み**
        # だった（#447。macro_data.created_at は一括収集日 07-08 に潰れており公表日で判定した）。
        "lag_days": 38,
        "stale_days": 62,
    },
]

# ── OECD SDMX API（sdmx.oecd.org・認証不要）────────────────────────────────
# 匿名クエリのみサポート・APIキー不要（OECD公式ドキュメントで確認済み・2026-07-09実API検証）。
# レート制限: 明確な閾値は非公開（"responsive experience維持のため導入"とのみ公式記載）。
# 月次nightly収集想定のため保守的に OECD_RATE_SLEEP=1.0s を挟む。
# 公表ラグ: CLI は毎月公表・対象月から2か月遅れ（例: 7月公表分は5月確定値。ただし直近1-2か月分は
# 暫定値として先行掲載されることがある）。先読みバイアス防止のため lag_days=60（e-Stat 鉱工業指数
# JP_IIP と同水準）。ADR-0009・Issue #283。
OECD_BASE_URL   = "https://sdmx.oecd.org/public/rest/data"
OECD_RATE_SLEEP = 1.0

OECD_SERIES: list[dict] = [
    {
        "code": "JP_CLI",
        "name": "日本 OECD景気先行指数（CLI・振幅調整済）",
        "category": "leading",
        # dataflow: エージェンシー,データセットID,バージョン。series_key: 9次元
        # REF_AREA.FREQ.MEASURE.UNIT_MEASURE.ACTIVITY.ADJUSTMENT.TRANSFORMATION.TIME_HORIZ.METHODOLOGY
        # を明示指定（空欄=ワイルドカードだとサーバー側デフォルト解決に依存するため不使用）。
        "dataflow":   "OECD.SDD.STES,DSD_STES@DF_CLI,4.1",
        "series_key": "JPN.M.LI.IX._Z.AA.IX._Z.H",
        "lag_days": 60,
        # stale_days=35: 健全時の理論最大 5日（観測周期30 + 実配信ラグ35 − lag_days 60）に
        # 観測周期30日を足した値＝「公表を1回スキップしたら鳴る」（ADR-0028）。lag_days が
        # 実配信ラグより大きいため freq 既定 105日との乖離が 21倍あり、既定のままでは OECD が
        # 配信を止めても100日気づけない（2026-08-04 実測 lag は 4日）。
        "stale_days": 35,
    },
]

# ── e-Stat API（CPI）──────────────────────────────────────────────────────────
# ESTAT_API_KEY が設定されている場合のみ収集（FRED_API_KEY と同挙動）。
# アカウント登録: https://www.e-stat.go.jp/api/ （無料・要ユーザー登録）
# GitHub Actions シークレット名: ESTAT_API_KEY
# statsDataId=0003427113: 2020年基準消費者物価指数（月次〜1970年・年次集計が同一テーブルに混在）
#   cdCat01=0001: 総合, 0161: 生鮮食品を除く総合（非季調）
#   cdArea=00000: 全国, 13A01: 東京都区部（表示名は "13100 東京都区部" だが実際の @code は 13A01。
#     旧来の "13100" を cdArea に指定すると STATUS=1「該当データなし」で0件になる・#262 で実API確認）
#   cdTab=1: 表章項目「指数」＋lvTime=4: 時間軸レベル「月次」（#262）。
#     両方指定しないと表章項目・時間軸レベルが絞られず年次系列のみ返却される。
ESTAT_API_KEY  = os.getenv("ESTAT_API_KEY", "")
ESTAT_BASE_URL = "https://api.e-stat.go.jp/rest/3.0/app/json/getStatsData"

#
# 公表日＝実配信日（e-Stat は公表と同時に反映される。2026年6月分の全国 CPI は 2026-07-24 に
# `macro_data.created_at` へ現れ、総務省の公表カレンダーと一致した）。anchor=period_start:
#   全国       翌月の「19日を含む週の金曜」8:30（週は日曜始まり＝最遅 25日）
#   東京都区部 当月の「26日を含む週の金曜」8:30（速報性が高く当月中に出る）
# lag_days は参照月の**期首**へ加算されるため、理論最大は「月の最大日数 31 + 公表日」で置く。
ESTAT_SERIES: list[dict] = [
    {
        "code": "JP_CPI_TOTAL",
        "name": "日本 CPI 全国総合",
        "category": "price",
        "stats_data_id": "0003427113",
        "cd_tab": "1",
        "cd_cat01": "0001",
        "cd_area": "00000",
        # 理論最大 56日＝31 + 25（19日を含む週の金曜の最遅）。実測 2026年6月分は 07-24 公表
        # ＝期首から 53日。旧値 30 は **23日の先読み**で、M-1/M-2/M-3/M-6 の学習・バックテスト
        # 全般を成績が良く見える方向へ歪めていた（#447）。
        "lag_days": 56,
        # stale_days=62: 理論最大 31日（観測周期31 + 実配信ラグ56 − lag_days 56）+ 観測周期31
        # ＝「公表を1回スキップしたら鳴る」（ADR-0028 規則5）。
        "stale_days": 62,
    },
    {
        "code": "JP_CPI_CORE",
        "name": "日本 CPI 全国コア（生鮮除く）",
        "category": "price",
        "stats_data_id": "0003427113",
        "cd_tab": "1",
        "cd_cat01": "0161",
        "cd_area": "00000",
        "lag_days": 56,   # 全国総合と同一リリース（同上）
        "stale_days": 62,
    },
    {
        "code": "JP_CPI_TOKYO",
        "name": "日本 CPI 東京都区部総合",
        "category": "price",
        "stats_data_id": "0003427113",
        "cd_tab": "1",
        "cd_cat01": "0001",
        "cd_area": "13A01",
        # 都区部は**当月分を当月中**に出す。26日が日曜だと金曜が月末（30日月なら翌月1日）に
        # ずれ、期首からの最遅が 30日＝旧値 30 はマージンがちょうど 0 だった。祝日ずれの 1 日を
        # 足して 31 とする（先読みは無かったので再収集は全国 CPI のついで・#447）。
        "lag_days": 31,
        "stale_days": 62,
    },
]

# ── e-Stat API（鉱工業指数・在庫指数）───────────────────────────────────────
# ESTAT_API_KEY を共用（CPI と同じキー・同じスキップ挙動）。CPI（ESTAT_SERIES）とは
# @time のフォーマットが異なる: CPI は "YYYY0000MM" の自己記述コードで直接パース可能だが、
# 鉱工業指数は "0500100" のような連番コードで年月を直接表現しない。そのため cd_tab/cd_area
# は存在せず（表章項目・地域軸を持たないテーブル）、fetch_estat_index_series が
# metaGetFlg="Y" でメタ情報（time 軸 code→"YYYYMM"）を同一レスポンスに同梱取得して変換する。
# 統計表は「経済産業省 鉱工業指数」2020年基準・業種別・季節調整済指数【月次】（2018年1月～）。
# 鉱工業指数は基準改定（2010→2015→2020年基準）のたびに statsDataId が別テーブルへ切り替わり
# 旧テーブルは更新停止する（FRED 版 JPNPROINDMISMEI が2024-04-30凍結した根本原因と同型・#253）。
# 次回基準改定時（目安10年ごと）は本節の statsDataId を再調査すること。
# cd_cat01="0001000" は業種分類（cat01）の「鉱工業総合」（"0002000"=製造工業も選択可）。
#
# **2026-08-04 時点で e-Stat 側が停止している（#451）**: 統計表 0004052177 / 0004052179 は
# `UPDATED_DATE=2026-06-03` を最後に更新されず、収録は2026年3月分（trade_date=2026-04-30）まで。
# 経産省サイトでは6月分が 2026-07-31 に公表済みなのに API へ来ない。実 API で全162表を確認した
# ところ世代は3つ（2026-06-03 / 2024-07-05 / 2021-09-08）しかなく**月次更新の痕跡が世代を通じて
# 存在しない**（世代間隔 約23か月）。収集側は無実で、API が返す 99 行を全て取れている。
# 昇格ゲート（`--features macro_jp_iip_yoy,macro_jp_iip_inventory_yoy`）でも4検定すべて非有意
# だったため既定特徴量からは棄却済み（`_GATE_REJECTED_FEATURES`）。**収集は継続する**——
# e-Stat がいずれ更新すればデータは貯まり、再判定できる。鮮度判定は `EXCLUDED_SERIES` で退避。
ESTAT_INDEX_SERIES: list[dict] = [
    {
        "code": "JP_IIP",
        "name": "日本 鉱工業生産指数（季調済・鉱工業総合）",
        "category": "real_economy",
        "stats_data_id": "0004052177",
        "cd_cat01": "0001000",
        "lag_days": 60,
        # stale_days=60: 理論最大 30日（観測周期30 + 実配信ラグ60 − lag_days 60）+ 観測周期30
        # ＝公表1回スキップで鳴る（ADR-0028）。現在は EXCLUDED_SERIES で判定から外しているが、
        # 除外を解いたときに freq 既定 105日では今回の 96日停止を検知できないため定義に残す。
        "stale_days": 60,
    },
    {
        "code": "JP_IIP_INVENTORY",
        "name": "日本 鉱工業在庫指数（季調済・鉱工業総合）",
        "category": "real_economy",
        "stats_data_id": "0004052179",
        "cd_cat01": "0001000",
        "lag_days": 60,
        "stale_days": 60,
    },
]

# ── 内閣府ESRI 直接CSV配布（GDP需要項目・認証不要）──────────────────────────
# e-Stat API は「四半期別ＧＤＰ速報」がvintage別アーカイブテーブル（四半期ごとに個別
# statsDataId）しか持たず継続更新系列が存在しないため不採用（#286 実地検証済み）。
# 代わりに内閣府ESRIが配布する実額系列CSVを直接取得する。1994年1-3月期〜最新四半期までの
# 連続時系列が単一ファイルに収録されているため、1回のfetchで全期間バックフィルできる。
# URL例: https://www.esri.cao.go.jp/jp/sna/data/data_list/sokuhou/files/2026/qe261_2/tables/gaku-jk2612.csv
#   ディレクトリ名 qe{yy}{q}_{report}・ファイル名 gaku-jk{yy}{q}{report}.csv
#   {yy}=西暦下2桁, {q}=四半期番号(1-4), {report}=速報回(1=1次速報,2=2次速報)
# 四半期ごとに新ディレクトリが発行され旧ディレクトリは(年次改定等で)消滅することがある
# （2026-07-10実API検証: 2024年分の旧1次速報ディレクトリは404）ため固定URLではなく
# プロービング方式で最新の有効URLを探す。HEAD/GETいずれも同一ステータスを返すことを
# 実地確認済みだが、ファイルが小さい（約40KB）ためGET一発でプローブ兼取得を行う。
ESRI_BASE_URL     = "https://www.esri.cao.go.jp/jp/sna/data/data_list/sokuhou/files"
ESRI_QUARTERS_BACK = 4      # 直近何四半期分を候補にするか（公表遅延・年次改定の保守的マージン）
ESRI_REPORTS       = (2, 1)  # 各四半期で試す速報回。2次速報（確定に近い）を優先、無ければ1次

# 単位: 十億円（2020年基準連鎖価格）・実質季節調整系列。
# lag_days: 1次速報は四半期末から約45〜50日で公表される（例: 1-3月期の1次速報は5月中旬）。
# trade_date は「参照四半期の翌四半期初日」を基準日とする（quarter_start + 3か月）ため、
# 四半期末からの経過日数とほぼ同じ意味になる。既存 JP_REAL_GDP（FRED版・lag_days=135）は
# FRED取り込み自体の遅延を含む値であり単純に流用できないため、実際の1次速報公表タイミング
# （45〜50日）に安全マージンを載せた60日を採用（e-Stat鉱工業指数 JP_IIP と同水準）。
ESRI_SERIES: list[dict] = [
    {
        "code": "JP_GDP_PRIVATE_CONSUMPTION",
        "name": "日本 GDP 民間最終消費支出",
        "category": "real_economy",
        "esri_column": "PrivateConsumption",
        "lag_days": 60,
    },
    {
        "code": "JP_GDP_RESIDENTIAL_INV",
        "name": "日本 GDP 民間住宅投資",
        "category": "real_economy",
        "esri_column": "PrivateResidentialInvestment",
        "lag_days": 60,
    },
    {
        "code": "JP_GDP_CAPEX",
        "name": "日本 GDP 民間企業設備投資",
        "category": "real_economy",
        "esri_column": "Private Non-Resi.Investment",
        "lag_days": 60,
    },
    {
        "code": "JP_GDP_PUBLIC_INV",
        "name": "日本 GDP 公的固定資本形成（公共投資）",
        "category": "real_economy",
        "esri_column": "PublicInvestment",
        "lag_days": 60,
    },
]

# ── 財務省「国債金利情報」CSV（認証不要・#458）──────────────────────────────
# **日次の日本10年金利**。M-3（ADR-0012）は週次高頻度ファクター専用だが、`dlm_jp10y` だけは
# 月次 `JP10Y_FRED` を週次差分へ落とす例外として grandfathered されていた（週次差分の 76.89%
# がゼロ＝#456 実測）。理由は「日次の日本10年金利ソースが無い」だったが、これは Yahoo/stooq が
# 全滅していたことを一般化した誤りで、**公的機関（財務省）が日次で直配信していた**。
#
# 実測（2026-08-07 再確認・#456 と一致）:
#   全期間版 jgbcm_all.csv  1,174,446B / cp932 / CRLF / データ13,270行・和暦パース失敗0
#                           10年物は 1986-07-05〜2026-07-31 の 9,909営業日（それ以前は "-"）
#                           **月次更新**（当月分は入らない）
#   当月版   jgbcm.csv      565B / 当月の営業日のみ（実測 R8.8.3〜8.5）＝**日次更新**
# したがって「初回は全期間版で埋め、以降は当月版で差分」の2本立てにする（下の need_full）。
#
# パースの注意点（すべて実測で確認済み）:
#   - 和暦（`R8.8.4` → 2026-08-04）。元号開始年は MOF_ERA_BASE
#   - cp932 / CRLF / ヘッダ2行（タイトル行＋列名行）/ 欠測は `-`
#   - 10年物は列 index 10（列名「10年」）。末尾に空行と注意書き行が付くため、
#     **先頭が元号記号の行だけをデータ行とみなす**
#
# ライセンス: PDL1.0（公共データ利用規約 第1.0版）＝出典表示のみで加工・再配布可。
# 制約が明示されているのは測量法に基づく地図類のみで数値データは対象外。
MOF_JGB_ALL_URL     = "https://www.mof.go.jp/jgbs/reference/interest_rate/data/jgbcm_all.csv"
MOF_JGB_CURRENT_URL = "https://www.mof.go.jp/jgbs/reference/interest_rate/jgbcm.csv"
MOF_ERA_BASE = {"M": 1867, "T": 1911, "S": 1925, "H": 1988, "R": 2018}

MOF_SERIES: list[dict] = [
    # lag_days=2: 基準日 T の値は **T+1 に公表**される（2026-08-07 JST 01:40 時点で最新が
    # 08-05＝08-06 分は未掲載＝当日公表ではない）。T+2 なら公表時刻に関わらず確実に既知。
    #
    # **日次系列では「安全側に大きく取る」ができない**（#447 と ADR-0028 の Consequences が
    # 低頻度系列で言っていることの裏返し）。lag_days が実配信ラグを超えると `trade_date` が
    # 未来日になり、鮮度ゲートは `ahead_days > 観測周期` を CRITICAL にする。日次の観測周期は
    # **1日**なので、2日以上の過大シフトは即 CRITICAL＝毎晩 exit 2 で自動起票が回り続ける
    # （実測: lag_days=4 で max(trade_date)=2026-08-09 ＝ today+2）。月次・四半期のように
    # 理論最大へ寄せる運用はここでは採れず、**上限は実配信ラグそのもの**になる。
    # 収集が数週間たまったら `macro_data.created_at` で実配信ラグを実測して引き直すこと
    # （ADR-0028 規則4/6。初回バックフィル行は created_at が潰れるので、バックフィル日より
    # 後に初出した観測だけを使う）。
    # freq="daily" 既定（許容14日）で足りる＝個別 `stale_days` は置かない（ADR-0028 の個別
    # 指定は lag_days が1周期を超える系列の話で、ここは超えない）。
    {"code": "JP10Y_MOF", "name": "日10年金利（財務省・日次）", "category": "rate",
     "mof_column": 10, "freq": "daily", "lag_days": 2},
]

# ── IMF WEO（World Economic Outlook）見通し（認証不要・#284）──────────────────
# 匿名クエリのみ・APIキー不要（2026-07-11実API検証済み）。既存チャネルは全て実績値
# （trailing）のため、予測・見通し（forward-looking）チャネルはこれが初。
#
# vintage（先読みバイアス）の扱い（重要・2系統構成）:
#   1) バックフィル: IMF公式「Historical WEO Forecasts Database」（WEOhistorical.xlsx・
#      https://www.imf.org/external/pubs/ft/weo/data/WEOhistorical.xlsx）。各vintage
#      （Spring/Fall・1990年〜）時点で「翌年」に対して発表されていた予測値を保持する
#      point-in-time パネル。2026-07-11取得版は Spring1990〜Fall2022 まで収録
#      （fetch_imf_weo_historical が実データで確認）。
#   2) 継続収集: 現行 SDMX API（api.imf.org/external/sdmx/3.0）の最新 dataflow（"+"）から
#      「今日時点で分かっている翌年予測値」を trade_date=収集日 で追加する（他の市場系列
#      と同じ「収集日に真に既知だった値」方式＝先読みバイアスなし）。
#   実証済みの注意点: 現行 dataflow は公式 vintage 境界とは無関係に**随時改定**される
#      （同一 COUNTRY_UPDATE_DATE 属性でも v6.0.0/v9.0.0 で値が異なることを確認）ため、
#      過去日付に遡って適用すると先読みバイアス化する＝③のような固定vintage日付割当は
#      使わない。
#   既知の空白: 2023年4月〜2025年4月の4vintage分はバックフィル source（WEOhistorical.xlsx）
#      にも公式vintage archive（IMF.RES配下の `WEO_2025_OCT_VINTAGE` 等・2025年10月開始の
#      新制度）にも収録されておらず復元不可。継続収集が始まった時点から新しいvintageが
#      発表されるたびに自然に埋まっていく。
IMF_HIST_URL = "https://www.imf.org/external/pubs/ft/weo/data/WEOhistorical.xlsx"
IMF_BASE_URL = "https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.RES/WEO/+"

IMF_SERIES: list[dict] = [
    {
        "code": "JP_WEO_GDP_FCAST",
        "name": "日本 IMF WEO 実質GDP成長率見通し（翌年）",
        "category": "forecast",
        "indicator":    "NGDP_RPCH",  # SDMX API 側の指標コード
        "excel_column": "ngdp_rpch",  # WEOhistorical.xlsx 側の列サフィックス
        "lag_days": 45,
    },
    {
        "code": "JP_WEO_CPI_FCAST",
        "name": "日本 IMF WEO インフレ率見通し（翌年）",
        "category": "forecast",
        "indicator":    "PCPIPCH",
        "excel_column": "pcpi_pch",
        "lag_days": 45,
    },
]

# ── GDELT DOC 2.0 API（api.gdeltproject.org・認証不要・#406）──────────────────
# 世界のニュース記事を横断集計するプロジェクト。`mode=timeline*` は日次の集計系列を
# JSON で返すため、**銘柄別ではなくマクロ集約系列**として macro_data へ入れる
# （銘柄別日次は 4,000社×250営業日 ≈ 370MB/年で Supabase 無料枠に入らない・#406）。
#
# 実API検証（2026-07-31）:
#   - 配信開始は 2017-01-01。それ以前を startdatetime に指定すると
#     "Invalid query start date"（GDELT_START が下限）。
#   - 2017-01-01〜今日を1リクエストで投げても間引かれず日次のまま返る（3,473点を確認）＝
#     系列あたり1リクエストで全履歴が揃う。ページングもチャンク分割も要らない。
#   - レート制限は「1リクエスト/5秒」（超過時は HTTP 200 で本文にプレーンテキストの
#     警告を返す＝JSON パース不能）。**実体は間隔だけでなく短時間の累積クエリ数にも効く**：
#     開発中に数十回叩いた直後は 60 秒バックオフしても全滅し、3 分放置すると即復帰した
#     （2026-07-31 実測）。よって n 回目の待ちを GDELT_RATE_SLEEP × n 秒とする線形バックオフ
#     （最大 6+12+18+24=60 秒/系列）で再試行し、それでも駄目なら graceful skip する。
#     通常運用（3系列 × 日1回）はこの上限に触れない。
GDELT_BASE_URL   = "https://api.gdeltproject.org/api/v2/doc/doc"
GDELT_RATE_SLEEP = 6.0
GDELT_RETRIES    = 4
GDELT_START      = "20170101000000"   # DOC 2.0 の配信開始（これ以前は Invalid query start date）

GDELT_SERIES: list[dict] = [
    # tone は正負を跨ぐ（悲観=負・楽観=正）ため yoy は不可＝zscore 規約（macro_snapshots 側）。
    {"code": "JP_NEWS_TONE",      "name": "日本ニュース 平均トーン（GDELT）",
     "category": "sentiment", "mode": "timelinetone", "query": "sourcecountry:japan"},
    {"code": "JP_NEWS_ECON_TONE", "name": "日本 株式市場ニュース 平均トーン（GDELT）",
     "category": "sentiment", "mode": "timelinetone",
     "query": "sourcecountry:japan theme:ECON_STOCKMARKET"},
    # timelinevol は「全記事に占める該当記事の割合(%)」＝報道量（関心度）。
    {"code": "JP_NEWS_ECON_VOL",  "name": "日本 株式市場ニュース 報道量（GDELT・全記事比%）",
     "category": "attention", "mode": "timelinevol",
     "query": "sourcecountry:japan theme:ECON_STOCKMARKET"},
]

# ── Wikimedia Pageviews API（wikimedia.org/api/rest_v1・認証不要・#406）────────
# ja.wikipedia の記事別日次閲覧数。**User-Agent に連絡先（URL かメール）を含めないと
# 403**（robot policy）＝ httpx 既定 UA では取得できない（2026-07-31 実API検証）。
# 配信開始は 2015-07-01（WIKIMEDIA_START）。1リクエストで全期間（4,047点）が返る。
#
# 単一記事はニュース以外の流入（リンク元の変化・編集）でも跳ねるため、テーマごとに
# **複数記事の合算バスケット**を1系列とする。欠測日（API が項目ごと落とす）は 0 ではなく
# 「その記事の寄与なし」として合算から除外する（記事追加/削除で水準が段差になるのを避ける）。
WIKIMEDIA_BASE_URL   = "https://wikimedia.org/api/rest_v1/metrics/pageviews/per-article"
WIKIMEDIA_UA         = "financial_app/1.0 (+https://github.com/kome-kome/financial_app)"
WIKIMEDIA_RATE_SLEEP = 1.0
WIKIMEDIA_START      = "20150701"   # Pageviews API の配信開始日
WIKIMEDIA_PROJECT    = "ja.wikipedia"

WIKIMEDIA_SERIES: list[dict] = [
    {"code": "JP_WIKI_MARKET_ATTN", "name": "日本 株式市場 関心度（ja.wikipedia 閲覧数）",
     "category": "attention", "articles": ["日経平均株価", "東京証券取引所"]},
    {"code": "JP_WIKI_MACRO_ATTN",  "name": "日本 景気・金融政策 関心度（ja.wikipedia 閲覧数）",
     "category": "attention",
     "articles": ["景気後退", "インフレーション", "日本銀行", "金融政策"]},
]


def _esri_candidate_urls(today: date) -> list[str]:
    """直近 ESRI_QUARTERS_BACK 四半期 × ESRI_REPORTS（2次優先）のURL候補を、
    新しい四半期・新しい速報回の順（最新四半期の2次速報が最優先）に生成する。"""
    q = (today.month - 1) // 3 + 1
    urls = []
    for i in range(ESRI_QUARTERS_BACK):
        total_q = today.year * 4 + (q - 1) - i
        year, qq = divmod(total_q, 4)
        qq += 1
        yy = f"{year % 100:02d}"
        for report in ESRI_REPORTS:
            urls.append(
                f"{ESRI_BASE_URL}/{year}/qe{yy}{qq}_{report}/tables/gaku-jk{yy}{qq}{report}.csv"
            )
    return urls


def _parse_esri_gdp_csv(text: str) -> dict[str, list[dict]]:
    """ESRI CSV本文（cp932デコード済みテキスト）を列名ごとの観測値へパースする。
    lag_days は未適用（trade_date は「参照四半期の翌四半期初日」の生値）。呼び出し側
    （fetch_esri_gdp_csv/collect_macro_data）が各系列の lag_days を適用する。
    CSVはカンマ区切り数値がダブルクォートで囲まれる箇所があるため csv モジュールでパースする
    （内部のカンマを区切りと誤認しない）。1列目は期ラベル（例 "1994/ 1- 3." "4- 6."）で、
    年は "/" 付きの行にのみ現れ、以降は前回の年を引き継ぐ。"""
    rows = list(csv.reader(io.StringIO(text)))

    header_idx = None
    for i, row in enumerate(rows):
        if any(cell.strip() == "PrivateConsumption" for cell in row):
            header_idx = i
            break
    if header_idx is None:
        log.warning("ESRI CSV ヘッダー行（PrivateConsumption）が見つからない")
        return {}

    header = [c.strip() for c in rows[header_idx]]
    col_index: dict[str, int] = {}
    for series in ESRI_SERIES:
        col_name = series["esri_column"]
        if col_name in col_index:
            continue
        try:
            col_index[col_name] = header.index(col_name)
        except ValueError:
            log.warning("ESRI CSV に列 %s が見つからない", col_name)
    if not col_index:
        return {}

    result: dict[str, list[dict]] = {name: [] for name in col_index}
    last_year: Optional[int] = None
    for row in rows[header_idx + 1:]:
        if not row or not row[0].strip():
            continue
        label = row[0].strip()
        if "/" in label:
            year_str, rest = label.split("/", 1)
            year_str = year_str.strip()
            if not year_str.isdigit():
                continue
            last_year = int(year_str)
        else:
            rest = label
        if last_year is None:
            continue
        month_str = rest.split("-", 1)[0].strip()
        if not month_str.isdigit():
            continue  # 脚注行（"＊年率で表示している。" 等）はここで除外される
        month = int(month_str)
        if month not in (1, 4, 7, 10):
            continue
        # 参照四半期の翌四半期初日を基準日とする（lag_days は呼び出し側で加算）。
        next_month = month + 3
        next_year  = last_year
        if next_month > 12:
            next_month -= 12
            next_year  += 1
        base_date = date(next_year, next_month, 1)

        for col_name, idx in col_index.items():
            if idx >= len(row):
                continue
            raw = row[idx].replace(",", "").strip()
            if not raw:
                continue
            try:
                value = float(raw)
            except ValueError:
                continue
            result[col_name].append({
                "trade_date": base_date.isoformat(),
                "open": None, "high": None, "low": None,
                "close": value, "volume": None,
            })
    return result


async def fetch_esri_gdp_csv(session: httpx.AsyncClient) -> dict[str, list[dict]]:
    """内閣府ESRI直接CSV配布からGDP需要項目CSVを1回のfetchで取得しパースする。
    直近 ESRI_QUARTERS_BACK 四半期×ESRI_REPORTS の候補URLを新しい順に試行し、
    最初にHTTP 200が返ったものを採用する（1ファイルに1994Q1〜最新まで全期間を含むため
    1回のfetchで足りる）。全候補が失敗した場合はログ警告のみで空dictを返す
    （他系列の収集を止めない・fetch_boj_series等と同じフェイルセーフ方針）。"""
    for url in _esri_candidate_urls(date.today()):
        try:
            r = await session.get(url, timeout=30)
        except Exception as e:
            log.debug("ESRI 取得失敗 %s: %s", url, type(e).__name__)
            continue
        if r.status_code != 200:
            continue
        try:
            text = r.content.decode("cp932")
        except UnicodeDecodeError as e:
            log.warning("ESRI CSV デコード失敗 %s: %s", url, e)
            return {}
        return _parse_esri_gdp_csv(text)

    log.warning(
        "ESRI GDP CSV 取得失敗: 候補URL全滅（直近%d四半期×速報%s）",
        ESRI_QUARTERS_BACK, ESRI_REPORTS,
    )
    return {}


def parse_mof_jgb_csv(text: str, column: int) -> list[dict]:
    """財務省「国債金利情報」CSV をパースして [{trade_date, close, ...}] を返す（#458）。

    text: cp932 からデコード済みの本文。column: 年限の列 index（10年物は 10）。
    ヘッダ2行・末尾の空行/注意書き行・欠測 `-` を落とし、和暦をISO日付へ変換する。
    **先頭が元号記号の行だけをデータ行とみなす**（注意書き行「最新の csv データが…」を
    行数や位置で切ると、行が増減したときに黙って壊れる）。
    """
    rows: list[dict] = []
    for cells in csv.reader(io.StringIO(text)):
        if not cells:
            continue
        head = cells[0].strip()
        if not head or head[0] not in MOF_ERA_BASE:
            continue          # タイトル行・列名行・空行・注意書き行
        try:
            era, rest = head[0], head[1:]
            y, m, d = rest.split(".")
            trade_date = date(MOF_ERA_BASE[era] + int(y), int(m), int(d))
        except (ValueError, KeyError):
            log.warning("MOF JGB: 和暦のパースに失敗しスキップ: %r", head)
            continue
        raw = cells[column].strip() if len(cells) > column else ""
        if not raw or raw == "-":
            continue          # 欠測（1986年以前の年限など）
        try:
            close = float(raw)
        except ValueError:
            log.warning("MOF JGB: 数値のパースに失敗しスキップ: %s %r", trade_date, raw)
            continue
        rows.append({"trade_date": trade_date.isoformat(), "close": close,
                     "open": None, "high": None, "low": None, "volume": None})
    return rows


async def fetch_mof_jgb_csv(session: httpx.AsyncClient, column: int,
                            full: bool) -> list[dict]:
    """財務省の国債金利 CSV を取得しパースする（#458）。

    full=True で全期間版（月次更新・1986年〜）、False で当月版（日次更新・当月のみ）。
    取得・パースの失敗はログ警告のみで空リストを返す（他系列の収集を止めない＝
    fetch_esri_gdp_csv / fetch_boj_series と同じフェイルセーフ方針）。
    """
    url = MOF_JGB_ALL_URL if full else MOF_JGB_CURRENT_URL
    try:
        r = await session.get(url, timeout=60)
        r.raise_for_status()
    except Exception as e:
        log.warning("MOF JGB 取得失敗 %s: %s", url, type(e).__name__)
        return []
    try:
        text = r.content.decode("cp932")
    except UnicodeDecodeError:
        log.warning("MOF JGB: cp932 でデコードできない %s", url)
        return []
    rows = parse_mof_jgb_csv(text, column)
    if not rows:
        log.warning("MOF JGB: データ行が1件も取れない %s", url)
    return rows


def _esri_apply_lag(rows: list[dict], lag_days: int) -> list[dict]:
    """_parse_esri_gdp_csv が返す生の観測値行に lag_days 分のシフトを適用する。"""
    if not lag_days:
        return rows
    shifted = []
    for r in rows:
        d = date.fromisoformat(r["trade_date"]) + timedelta(days=lag_days)
        shifted.append({**r, "trade_date": d.isoformat()})
    return shifted


async def fetch_fred_series(
    session: httpx.AsyncClient,
    fred_id: str,
    date_from: str,  # "YYYY-MM-DD"
    date_to:   str,  # "YYYY-MM-DD"
    lag_days: int = 0,
) -> list:
    """FRED API から指定系列の観測値を取得する（日次・月次両対応）。
    欠損値（"."）と None はスキップ。月次系列は FRED が1か月1観測を返すので結果も月次になる。
    lag_days > 0 のとき、observation の日付（期の参照開始日）を lag_days 分だけ後ろへ
    シフトして trade_date とする＝公表ラグ補正（実体経済指標の先読みバイアス防止）。"""
    params = {
        "series_id":         fred_id,
        "api_key":           FRED_API_KEY,
        "file_type":         "json",
        "observation_start": date_from,
        "observation_end":   date_to,
    }
    try:
        r = await session.get(FRED_BASE_URL, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
    except httpx.HTTPStatusError as e:
        # URL にクエリで api_key を渡しているため e をそのまま出すと鍵が漏洩する
        log.warning("FRED 取得失敗 %s: HTTP %s", fred_id, e.response.status_code)
        return []
    except Exception as e:
        log.warning("FRED 取得失敗 %s: %s", fred_id, type(e).__name__)
        return []

    rows = []
    for obs in data.get("observations", []):
        v = obs.get("value", ".")
        if v == "." or v is None:
            continue
        try:
            obs_date = obs["date"]
            if lag_days:
                obs_date = (date.fromisoformat(obs_date) + timedelta(days=lag_days)).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open":   None,
                "high":   None,
                "low":    None,
                "close":  float(v),
                "volume": None,
            })
        except (ValueError, KeyError):
            continue
    return rows


async def fetch_boj_series(
    session: httpx.AsyncClient,
    db: str,
    boj_code: str,
    date_from: str,  # "YYYYMM"
    date_to:   str,  # "YYYYMM"
    lag_days: int = 0,
    freq: str = "monthly",
) -> list:
    """日銀時系列統計 API（stat-search.boj.or.jp/api/v1/getDataCode）から観測値を取得する。
    monthly: SURVEY_DATES は YYYYMM。quarterly: SURVEY_DATES は YYYYQQ（01-04=Q1-Q4）。
    四半期 Q1=4月公表, Q2=7月公表, Q3=10月公表, Q4=翌年1月公表 として calendar date へ変換後
    lag_days 分だけ後ろへシフトして trade_date とする。
    quarterly 系列の startDate/endDate は YYYYQQ 形式に変換して送信（YYYYMM だと 400）。"""
    _Q_RELEASE_MONTH = {1: 4, 2: 7, 3: 10, 4: 1}

    if freq == "quarterly":
        def _yyyymm_to_boj_quarter(yyyymm: str) -> str:
            year, month = int(yyyymm[:4]), int(yyyymm[4:])
            if month <= 3:   return f"{year - 1}04"  # Jan-Mar → Q4 of prev year
            elif month <= 6: return f"{year}01"       # Apr-Jun → Q1
            elif month <= 9: return f"{year}02"       # Jul-Sep → Q2
            else:            return f"{year}03"        # Oct-Dec → Q3
        date_from = _yyyymm_to_boj_quarter(date_from)
        date_to   = _yyyymm_to_boj_quarter(date_to)

    params = {
        "format":    "json",
        "db":        db,
        "startDate": date_from,
        "endDate":   date_to,
        "code":      boj_code,
    }
    try:
        r = await session.get(f"{BOJ_BASE_URL}/getDataCode", params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.warning("BOJ 取得失敗 %s/%s: %s", db, boj_code, type(e).__name__)
        return []

    if data.get("STATUS") != 200:
        log.warning("BOJ 取得失敗 %s/%s: STATUS=%s", db, boj_code, data.get("STATUS"))
        return []

    rows = []
    for series in data.get("RESULTSET", []):
        vdata = series.get("VALUES", {})
        survey_dates = vdata.get("SURVEY_DATES", [])
        values       = vdata.get("VALUES", [])
        for sd, v in zip(survey_dates, values):
            if v is None:
                continue
            sd_str = str(sd)
            if freq == "quarterly":
                year    = int(sd_str[:4])
                quarter = int(sd_str[4:])
                month   = _Q_RELEASE_MONTH[quarter]
                if quarter == 4:
                    year += 1
                obs_date = date(year, month, 1).isoformat()
            else:
                year     = int(sd_str[:4])
                month    = int(sd_str[4:])
                obs_date = date(year, month, 1).isoformat()
            if lag_days:
                obs_date = (date.fromisoformat(obs_date) + timedelta(days=lag_days)).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open": None, "high": None, "low": None,
                "close": float(v), "volume": None,
            })
    return rows


async def fetch_oecd_series(
    session: httpx.AsyncClient,
    dataflow: str,
    series_key: str,
    date_from: str,  # "YYYY-MM"
    lag_days: int = 0,
) -> list:
    """OECD SDMX API（sdmx.oecd.org/public/rest/data）から月次系列を取得する。
    認証不要（匿名クエリのみサポート・APIキー不要）。CSV形式（csvfilewithlabels）で取得し
    TIME_PERIOD（"YYYY-MM"）/OBS_VALUE 列をパースする。存在しない series_key は
    404 "NoRecordsFound" を返す（2026-07-09実API検証済み）。"""
    url = f"{OECD_BASE_URL}/{dataflow}/{series_key}"
    params = {"startPeriod": date_from, "format": "csvfilewithlabels"}
    try:
        r = await session.get(url, params=params, timeout=30)
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text), usecols=["TIME_PERIOD", "OBS_VALUE"])
    except Exception as e:
        log.warning("OECD 取得失敗 %s: %s", series_key, type(e).__name__)
        return []

    rows = []
    for _, row in df.dropna(subset=["OBS_VALUE"]).iterrows():
        try:
            year, month = (int(x) for x in str(row["TIME_PERIOD"]).split("-"))
            obs_date = date(year, month, 1).isoformat()
            if lag_days:
                obs_date = (date.fromisoformat(obs_date) + timedelta(days=lag_days)).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open": None, "high": None, "low": None,
                "close": float(row["OBS_VALUE"]), "volume": None,
            })
        except (ValueError, TypeError):
            continue
    return rows


async def fetch_estat_series(
    session: httpx.AsyncClient,
    stats_data_id: str,
    cd_tab: str,
    cd_cat01: str,
    cd_area: str,
    date_from: str,  # "YYYYMM000000"
    date_to:   str,  # "YYYYMM000000"
    lag_days: int = 0,
) -> list:
    """e-Stat API（api.e-stat.go.jp）から月次統計を取得する。ESTAT_API_KEY が必要。
    @time 実測フォーマットは月次 "YYYY" + "00" + "MM" + "MM"（月を2回繰り返す。例 2024年12月＝
    "2024001212"）・年度（会計年度集計）は "YYYY" + "10" + "0000"。年度行の先頭6文字が偶然
    "YYYY10" になり月=10と誤読される事故があった（#256）ため、月は末尾2文字から取り出す。
    cdTab（表章項目=1:指数）と lvTime（時間軸レベル=4:月次）の両方が必須（#262 で実 API 検証済み。
    片方だけでは年次行が混入するか解析失敗になる。過去の lvTime 単体試行が失敗したのは cdTab
    未指定のままだったため）。"""
    params = {
        "appId":        ESTAT_API_KEY,
        "statsDataId":  stats_data_id,
        "cdTab":        cd_tab,
        "cdCat01":      cd_cat01,
        "cdArea":       cd_area,
        "cdTimeFrom":   date_from,
        "cdTimeTo":     date_to,
        "lvTime":       "4",
        "lang":         "J",
        "metaGetFlg":   "N",
    }
    try:
        r = await session.get(ESTAT_BASE_URL, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.warning("e-Stat 取得失敗 %s/%s: %s", stats_data_id, cd_cat01, type(e).__name__)
        return []

    try:
        values = (
            data["GET_STATS_DATA"]["STATISTICAL_DATA"]["DATA_INF"]["VALUE"]
        )
    except (KeyError, TypeError):
        log.warning("e-Stat レスポンス解析失敗 %s/%s", stats_data_id, cd_cat01)
        return []

    if isinstance(values, dict):
        values = [values]

    rows = []
    for val in values:
        raw_v = val.get("$")
        t     = val.get("@time", "")
        if raw_v is None or t == "":
            continue
        # @cat01/@area が VALUE 要素に存在する場合のみフィルタ（属性なし = API 側で既に絞込済み）。
        # 属性なしのとき None != cd_cat01 → 全行スキップになるため "in val" チェックが必須。
        if cd_cat01 and "@cat01" in val and val["@cat01"] != cd_cat01:
            continue
        if cd_area and "@area" in val and val["@area"] != cd_area:
            continue
        try:
            # "YYYY" + "00" + "MM" + "MM"（月次・月が2回繰り返される）。年は先頭4文字、
            # 月は末尾2文字（[6:8] と同値）から取り出す。[4:6] は月ではなく年次/月次の
            # 区分マーカー（月次="00"・年度="10"）なので月として読んではいけない。
            year   = int(t[:4])
            month  = int(t[8:10])
            obs_date = date(year, month, 1).isoformat()
            if lag_days:
                obs_date = (date.fromisoformat(obs_date) + timedelta(days=lag_days)).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open": None, "high": None, "low": None,
                "close": float(raw_v), "volume": None,
            })
        except (ValueError, IndexError):
            continue
    # 同 trade_date が複数行ある場合（API が同一時点を複数カテゴリで返す等）は最後の値で dedup。
    seen: dict = {}
    for r in rows:
        seen[r["trade_date"]] = r
    return list(seen.values())


async def fetch_estat_index_series(
    session: httpx.AsyncClient,
    stats_data_id: str,
    cd_cat01: str,
    lag_days: int = 0,
) -> list:
    """e-Stat API から「time 軸が連番コード」形式の指数系列（鉱工業指数等）を取得する。
    CPI 系列（fetch_estat_series）は @time が "YYYY0000MM" の自己記述コードで直接パース
    できるが、鉱工業指数は @time が "0500100" のような連番コードで年月を直接表現しない。
    metaGetFlg="Y" を付けて time 軸のメタ情報（code→"YYYYMM"）を同一レスポンスへ同梱させ、
    そのマッピングで変換する（追加の getMetaInfo 呼び出し不要・1リクエストで完結）。
    ウエイト行等（"付加生産ウエイト" 等・6桁の YYYYMM にならない）はスキップする。"""
    params = {
        "appId":       ESTAT_API_KEY,
        "statsDataId": stats_data_id,
        "cdCat01":     cd_cat01,
        "metaGetFlg":  "Y",
        "lang":        "J",
    }
    try:
        r = await session.get(ESTAT_BASE_URL, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.warning("e-Stat(index) 取得失敗 %s/%s: %s", stats_data_id, cd_cat01, type(e).__name__)
        return []

    try:
        stat       = data["GET_STATS_DATA"]["STATISTICAL_DATA"]
        values     = stat["DATA_INF"]["VALUE"]
        class_objs = stat["CLASS_INF"]["CLASS_OBJ"]
    except (KeyError, TypeError):
        log.warning("e-Stat(index) レスポンス解析失敗 %s/%s", stats_data_id, cd_cat01)
        return []

    if isinstance(values, dict):
        values = [values]
    if isinstance(class_objs, dict):
        class_objs = [class_objs]

    time_map: dict = {}
    for obj in class_objs:
        if obj.get("@id") != "time":
            continue
        classes = obj.get("CLASS", [])
        if isinstance(classes, dict):
            classes = [classes]
        for c in classes:
            time_map[c.get("@code")] = c.get("@name")

    rows = []
    for val in values:
        raw_v  = val.get("$")
        code   = val.get("@time", "")
        yyyymm = time_map.get(code, "")
        if raw_v is None or len(yyyymm) != 6 or not yyyymm.isdigit():
            continue  # ウエイト行等（time_map の名前が YYYYMM でない）はスキップ
        try:
            year, month = int(yyyymm[:4]), int(yyyymm[4:])
            obs_date = date(year, month, 1).isoformat()
            if lag_days:
                obs_date = (date.fromisoformat(obs_date) + timedelta(days=lag_days)).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open": None, "high": None, "low": None,
                "close": float(raw_v), "volume": None,
            })
        except (ValueError, IndexError):
            continue
    return rows


def _parse_imf_weo_sheet(wb, excel_column: str, lag_days: int = 0) -> list[dict]:
    """WEOhistorical.xlsx の1シート（例 "ngdp_rpch"）から日本（JPN）の「翌年予測値」を
    vintage（S{year}=Spring/F{year}=Fall）ごとに抽出する。シートは1行=対象年（`year`列）、
    列は vintage ごとの `S{year}{excel_column}`/`F{year}{excel_column}` ペア。vintage 年 Y の
    予測値は「対象年=Y+1の行」から読み取る（＝当年ではなく翌年見通しを採用し、既に判明
    済みの当年実績と区別する）。欠測は "." 文字列（IMF側の欠測記号）。"""
    ws = wb[excel_column]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    col_index = {name: i for i, name in enumerate(header) if name}
    iso_idx  = col_index.get("ISOAlpha_3Code")
    year_idx = col_index.get("year")
    if iso_idx is None or year_idx is None:
        log.warning("IMF WEO historical シート %s: 想定列（ISOAlpha_3Code/year）が見つからない", excel_column)
        return []

    by_year: dict[int, tuple] = {}
    for row in rows_iter:
        if row[iso_idx] != "JPN":
            continue
        try:
            by_year[int(row[year_idx])] = row
        except (TypeError, ValueError):
            continue

    vintage_years = sorted({
        int(name[1:5])
        for name in col_index
        if name[:1] in ("S", "F") and name[1:5].isdigit() and name[5:] == excel_column
    })

    result = []
    for vy in vintage_years:
        target_row = by_year.get(vy + 1)
        if target_row is None:
            continue
        for prefix, month in (("S", 4), ("F", 10)):
            ci = col_index.get(f"{prefix}{vy}{excel_column}")
            if ci is None or ci >= len(target_row):
                continue
            raw = target_row[ci]
            try:
                value = float(raw)  # 欠測は "." 文字列（IMF側の欠測記号）→ ValueError で除外
            except (TypeError, ValueError):
                continue
            base_date = date(vy, month, 1)
            trade_date = (base_date + timedelta(days=lag_days)) if lag_days else base_date
            result.append({
                "trade_date": trade_date.isoformat(),
                "open": None, "high": None, "low": None,
                "close": value, "volume": None,
            })
    return result


async def fetch_imf_weo_historical(session: httpx.AsyncClient) -> dict[str, list[dict]]:
    """IMF公式「Historical WEO Forecasts Database」（WEOhistorical.xlsx）から
    IMF_SERIES 全系列の point-in-time バックフィルデータを1回のfetchで取得する
    （ESRI GDP CSV と同型：1ファイルに全vintage×全年が入っているため1回で足りる）。
    サーバーは User-Agent のみの素の GET を 403 で拒否するが、Range ヘッダーを付けた
    リクエストには 200/206 で応答する（bot対策の実装差・2026-07-11実API検証済み）。
    失敗時は空dictを返し、他系列の収集を止めない（fetch_esri_gdp_csv と同じ方針）。"""
    try:
        r = await session.get(IMF_HIST_URL, headers={"Range": "bytes=0-"}, timeout=120)
        if r.status_code not in (200, 206):
            log.warning("IMF WEO historical 取得失敗: HTTP %s", r.status_code)
            return {}
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    except Exception as e:
        log.warning("IMF WEO historical 取得/読込失敗: %s", type(e).__name__)
        return {}

    result: dict[str, list[dict]] = {}
    try:
        for series in IMF_SERIES:
            sheet_name = series["excel_column"]
            if sheet_name not in wb.sheetnames:
                log.warning("IMF WEO historical シート不在: %s", sheet_name)
                continue
            result[series["code"]] = _parse_imf_weo_sheet(wb, sheet_name, series.get("lag_days", 0))
    finally:
        wb.close()
    return result


async def fetch_imf_weo_current(session: httpx.AsyncClient, indicator: str) -> list:
    """現行（最新）IMF WEO dataflow（api.imf.org・認証不要）から「収集日時点で分かって
    いる翌年予測値」を1点取得する。trade_date は収集日そのもの（他の市場系列と同じ
    「その日に真に既知だった値」方式＝先読みバイアスなし）。現行dataflowは公式vintage
    境界と無関係に随時改定される（同一 COUNTRY_UPDATE_DATE 属性でも過去版と値が異なる
    ことを実API確認済み）ため、過去日付への割当（バックフィルと同じlag_days方式）はせず
    常に「今日」に紐づける。"""
    today = date.today()
    target_year = str(today.year + 1)
    url = f"{IMF_BASE_URL}/JPN.{indicator}.A"
    try:
        r = await session.get(
            url,
            params={"dimensionAtObservation": "TIME_PERIOD", "attributes": "dsd", "measures": "all"},
            headers={"Accept": "application/vnd.sdmx.data+json; version=2.0.0"},
            timeout=30,
        )
        r.raise_for_status()
        d = r.json()
        struct = d["data"]["structures"][0]
        tp = [v["value"] for dim in struct["dimensions"]["observation"] if dim["id"] == "TIME_PERIOD" for v in dim["values"]]
        series = list(d["data"]["dataSets"][0]["series"].values())[0]
        obs = {tp[int(k)]: v[0] for k, v in series["observations"].items()}
    except Exception as e:
        log.warning("IMF WEO current 取得失敗 %s: %s", indicator, type(e).__name__)
        return []

    raw = obs.get(target_year)
    if raw is None:
        return []
    try:
        value = float(raw)
    except (TypeError, ValueError):
        return []
    return [{
        "trade_date": today.isoformat(),
        "open": None, "high": None, "low": None,
        "close": value, "volume": None,
    }]


async def fetch_gdelt_timeline(
    session: httpx.AsyncClient,
    query: str,
    mode: str,          # "timelinetone" | "timelinevol"
    date_from: str,     # "YYYYMMDDHHMMSS"
    date_to: str,       # "YYYYMMDDHHMMSS"
) -> list:
    """GDELT DOC 2.0 API から日次の集計系列（トーン／報道量）を取得する。認証不要。

    レート超過時は HTTP 200 のままプレーンテキストの警告本文を返すため、**ステータス
    コードではなく本文が JSON かどうか**で判定し、GDELT_RATE_SLEEP × 試行回数（線形
    バックオフ）だけ待って再試行する。取得できなければ空リスト（graceful skip・他コネクタ
    と同型）＝その系列はこの回だけ欠測し、次回収集で埋まる。"""
    params = {"query": query, "mode": mode, "format": "json",
              "startdatetime": date_from, "enddatetime": date_to}
    payload = None
    for attempt in range(GDELT_RETRIES):
        try:
            r = await session.get(GDELT_BASE_URL, params=params, timeout=180)
            r.raise_for_status()
            body = r.text.lstrip()
            if body.startswith("{"):
                payload = r.json()
                break
            log.debug("GDELT レート制限 %s (%d回目): %s", mode, attempt + 1, body[:80])
        except Exception as e:
            log.debug("GDELT 取得失敗 %s (%d回目): %s", mode, attempt + 1, type(e).__name__)
        await asyncio.sleep(GDELT_RATE_SLEEP * (attempt + 1))

    if payload is None:
        log.warning("GDELT 取得失敗 %s %s: リトライ上限", mode, query)
        return []

    timeline = payload.get("timeline") or []
    if not timeline:
        return []

    rows = []
    for point in timeline[0].get("data") or []:
        try:
            stamp = str(point["date"])[:8]          # "20240101T000000Z" → "20240101"
            obs_date = date(int(stamp[:4]), int(stamp[4:6]), int(stamp[6:8])).isoformat()
            rows.append({
                "trade_date": obs_date,
                "open": None, "high": None, "low": None,
                "close": float(point["value"]), "volume": None,
            })
        except (KeyError, TypeError, ValueError):
            continue
    return rows


async def fetch_wikimedia_pageviews(
    session: httpx.AsyncClient,
    articles: list[str],
    date_from: str,   # "YYYYMMDD"
    date_to: str,     # "YYYYMMDD"
) -> list:
    """Wikimedia Pageviews API から記事バスケットの日次閲覧数合計を取得する。認証不要。

    記事ごとに1リクエスト（全期間が1回で返る）し、日付キーで合算する。存在しない記事は
    404 を返すので、その記事だけ落として残りで合算を続ける（graceful skip）。欠測日は
    0 埋めせず合算から除外する（記事の増減で水準に段差を作らないため）。"""
    totals: dict[str, float] = {}
    for i, article in enumerate(articles):
        quoted = urlquote(article, safe="")
        url = (f"{WIKIMEDIA_BASE_URL}/{WIKIMEDIA_PROJECT}/all-access/all-agents/"
               f"{quoted}/daily/{date_from}/{date_to}")
        try:
            r = await session.get(url, headers={"User-Agent": WIKIMEDIA_UA}, timeout=60)
            r.raise_for_status()
            items = r.json().get("items") or []
        except Exception as e:
            log.warning("Wikimedia 取得失敗 %s: %s", article, type(e).__name__)
            items = []
        for item in items:
            try:
                stamp = str(item["timestamp"])[:8]   # "2024010100" → "20240101"
                key = date(int(stamp[:4]), int(stamp[4:6]), int(stamp[6:8])).isoformat()
                totals[key] = totals.get(key, 0.0) + float(item["views"])
            except (KeyError, TypeError, ValueError):
                continue
        if i < len(articles) - 1:
            await asyncio.sleep(WIKIMEDIA_RATE_SLEEP)

    return [{
        "trade_date": d,
        "open": None, "high": None, "low": None,
        "close": v, "volume": None,
    } for d, v in sorted(totals.items())]


async def fetch_yahoo_history(
    session: httpx.AsyncClient,
    yf_ticker: str,
    date_from: str,   # "YYYYMMDD"
    date_to:   str,   # "YYYYMMDD"
) -> list:
    """Yahoo Finance v8 API から日次 OHLCV を取得する。
    GitHub Actions（Azure IP）からも動作する。stooq の代替として使用。"""
    try:
        # date → Unix timestamp（JST 00:00 = UTC 前日15:00、余裕を持って+1日）
        y1, m1, d1_ = int(date_from[:4]), int(date_from[4:6]), int(date_from[6:8])
        y2, m2, d2_ = int(date_to[:4]),   int(date_to[4:6]),   int(date_to[6:8])
        period1 = int(calendar.timegm((y1, m1, d1_, 0, 0, 0)))
        period2 = int(calendar.timegm((y2, m2, d2_, 23, 59, 59)))
    except (ValueError, IndexError) as e:
        log.debug(f"Yahoo Finance 日付変換失敗 {yf_ticker}: {e}")
        return []

    url = (f"https://query1.finance.yahoo.com/v8/finance/chart/{yf_ticker}"
           f"?interval=1d&period1={period1}&period2={period2}")
    try:
        r = await session.get(url, timeout=30,
                              headers={"User-Agent": "Mozilla/5.0",
                                       "Accept": "application/json"})
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        log.debug(f"Yahoo Finance 取得失敗 {yf_ticker}: {e}")
        return []

    try:
        result = data["chart"]["result"][0]
        timestamps = result["timestamp"]
        quote = result["indicators"]["quote"][0]
    except (KeyError, IndexError, TypeError) as e:
        log.debug(f"Yahoo Finance レスポンス解析失敗 {yf_ticker}: {e}")
        return []

    def _sf(lst, i):
        v = lst[i] if i < len(lst) else None
        return float(v) if v is not None else None

    rows = []
    for i, ts in enumerate(timestamps):
        close = _sf(quote.get("close", []), i)
        if close is None:
            continue
        rows.append({
            "trade_date": date.fromtimestamp(ts).strftime("%Y-%m-%d"),
            "open":   _sf(quote.get("open",   []), i),
            "high":   _sf(quote.get("high",   []), i),
            "low":    _sf(quote.get("low",    []), i),
            "close":  close,
            "volume": _sf(quote.get("volume", []), i),
        })
    return rows


async def fetch_stooq_history(
    session: httpx.AsyncClient,
    ticker:  str,
    date_from: str,   # "YYYYMMDD"
    date_to:   str,   # "YYYYMMDD"
) -> list:
    """stooq 日次 OHLCV（汎用ティッカー・マクロ用）。open/high/low/volume は None 許容。"""
    return await _fetch_stooq_ohlcv(
        session, ticker, date_from, date_to,
        strict=False, log_label=f"stooq マクロ取得失敗 {ticker}",
    )


async def collect_macro_data(
    db,
    years_back: int = 5,
    on_progress: Optional[Callable[[int, int, str], None]] = None,
    cancel_check: Optional[Callable[[], bool]] = None,
    only: Optional[list[str]] = None,
):
    """MACRO_SERIES（Yahoo/stooq）+ FRED_SERIES + BOJ_SERIES + OECD_SERIES + ESRI_SERIES +
    IMF_SERIES + ESTAT_SERIES + GDELT_SERIES + WIKIMEDIA_SERIES + MOF_SERIES を macro_data
    に upsert。
    Yahoo Finance 優先（GitHub Actions Azure IP 対応）→ stooq フォールバック。FRED:
    FRED_API_KEY 設定時のみ。BOJ・OECD・ESRI・IMF・GDELT・Wikimedia・MOF: 常時収集（認証不要）。
    e-Stat: ESTAT_API_KEY 設定時のみ。既存レコードは close 等を上書き（最新値で更新）。

    only: series_code のリスト。指定時はグループ横断で対象系列だけを収集する（#444）。
    1 系列の定義是正（freq/lag_days の修正など）に伴う再収集で全外部 API を叩き直す
    副作用——特に GDELT の累積クエリ制限の消費——を避けるための絞り込み。"""
    # グループ単位で先に絞る。ESRI/IMF は 1 リクエストで全系列ぶんを取るため、
    # ループ内スキップだけでは重い fetch が残る＝リストを空にして丸ごと回避する。
    only_codes = set(only) if only else None

    def _sel(series_list: list[dict]) -> list[dict]:
        """only 指定時は対象 code だけへ絞る（未指定ならそのまま全件）。"""
        if only_codes is None:
            return series_list
        return [s for s in series_list if s["code"] in only_codes]

    macro_list     = _sel(MACRO_SERIES)
    fred_list      = _sel(FRED_SERIES)
    boj_list       = _sel(BOJ_SERIES)
    oecd_list      = _sel(OECD_SERIES)
    esri_list      = _sel(ESRI_SERIES)
    imf_list       = _sel(IMF_SERIES)
    estat_list     = _sel(ESTAT_SERIES)
    estat_idx_list = _sel(ESTAT_INDEX_SERIES)
    gdelt_list     = _sel(GDELT_SERIES)
    wiki_list      = _sel(WIKIMEDIA_SERIES)
    mof_list       = _sel(MOF_SERIES)
    if only_codes:
        known = {s["code"] for s in (
            MACRO_SERIES + FRED_SERIES + BOJ_SERIES + OECD_SERIES + ESRI_SERIES
            + IMF_SERIES + ESTAT_SERIES + ESTAT_INDEX_SERIES + GDELT_SERIES + WIKIMEDIA_SERIES
            + MOF_SERIES
        )}
        unknown = sorted(only_codes - known)
        if unknown:
            # 打ち間違いを「0 件収集・成功」で黙って返さない（fail fast）。
            raise ValueError(f"未知の series_code: {', '.join(unknown)}")

    today      = date.today()
    start      = today - timedelta(days=int(years_back * 365.25))
    d1         = start.strftime("%Y%m%d")
    d2         = today.strftime("%Y%m%d")
    # FRED は市場系より長く遡る（四半期系列の zscore に ≥20 点を確保）。
    fred_start = today - timedelta(days=int(max(years_back, FRED_MIN_YEARS_BACK) * 365.25))
    d1_iso     = fred_start.strftime("%Y-%m-%d")
    d2_iso     = today.strftime("%Y-%m-%d")
    # BOJ: 短観は四半期なので FRED と同じく長めに遡る（zscore ≥20 点確保）。
    boj_start  = today - timedelta(days=int(max(years_back, FRED_MIN_YEARS_BACK) * 365.25))
    d1_boj     = boj_start.strftime("%Y%m")   # "YYYYMM"
    d2_boj     = today.strftime("%Y%m")
    # e-Stat: @time フォーマット YYYYMM000000。
    d1_estat   = boj_start.strftime("%Y%m") + "000000"
    d2_estat   = today.strftime("%Y%m") + "000000"
    # OECD: startPeriod は "YYYY-MM"。FRED/BOJ 同様に長めに遡る（zscore ≥20 点確保）。
    oecd_start = today - timedelta(days=int(max(years_back, FRED_MIN_YEARS_BACK) * 365.25))
    d1_oecd    = oecd_start.strftime("%Y-%m")
    # GDELT / Wikimedia（#406）: FRED と同じ窓で遡りつつ、各ソースの配信開始日で下限を切る
    # （それ以前を要求すると GDELT は "Invalid query start date" を返す）。
    d1_gdelt   = max(fred_start.strftime("%Y%m%d") + "000000", GDELT_START)
    d2_gdelt   = today.strftime("%Y%m%d") + "000000"
    d1_wiki    = max(fred_start.strftime("%Y%m%d"), WIKIMEDIA_START)
    d2_wiki    = today.strftime("%Y%m%d")
    total      = (
        len(macro_list)
        + (len(fred_list)  if FRED_API_KEY  else 0)
        + len(boj_list)
        + len(oecd_list)
        + len(esri_list)
        + len(imf_list)
        + (len(estat_list) + len(estat_idx_list) if ESTAT_API_KEY else 0)
        + len(gdelt_list)
        + len(wiki_list)
        + len(mof_list)
    )
    saved      = 0

    async with httpx.AsyncClient(timeout=60) as session:
        for i, series in enumerate(macro_list, 1):
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(i-1, total, "[マクロ収集] ユーザー停止")
                return saved

            # Yahoo Finance 優先（GitHub Actions Azure IP 対応）→ stooq フォールバック
            rows = await fetch_yahoo_history(session, series["yf_ticker"], d1, d2)
            src = "Yahoo Finance"
            if not rows:
                rows = await fetch_stooq_history(session, series["ticker"], d1, d2)
                src = "stooq"
            if on_progress:
                on_progress(i-1, total, f"[マクロ {i}/{total}] {series['name']} ({src}) 取得中")
            if not rows:
                if on_progress:
                    on_progress(i, total, f"[マクロ {i}/{total}] {series['name']} データ無し")
                continue

            # 系列単位のバルク upsert（(series_code, trade_date) 競合で最新値上書き）。
            # 旧実装は行ごとに INSERT/UPDATE を発行していたが N+1 解消のため 1 文に圧縮。
            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(i, total, f"[マクロ {i}/{total}] {series['name']}: {n}件処理")

        # ── FRED 収集（FRED_API_KEY が設定されている場合のみ）──────────────────
        if not FRED_API_KEY:
            if on_progress:
                on_progress(len(macro_list), total, "[FRED] FRED_API_KEY 未設定のためスキップ")
        else:
            base_i = len(macro_list)
            for j, series in enumerate(fred_list, 1):
                idx = base_i + j
                if cancel_check and cancel_check():
                    if on_progress:
                        on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                    return saved

                if on_progress:
                    on_progress(idx - 1, total, f"[FRED {j}/{len(fred_list)}] {series['name']} 取得中")
                rows = await fetch_fred_series(
                    session, series["fred_id"], d1_iso, d2_iso, series.get("lag_days", 0)
                )
                await asyncio.sleep(FRED_RATE_SLEEP)

                if not rows:
                    if on_progress:
                        on_progress(idx, total, f"[FRED {j}/{len(fred_list)}] {series['name']} データ無し")
                    continue

                vals = [{
                    "series_code": series["code"],
                    "series_name": series["name"],
                    "category":    series["category"],
                    "trade_date":  r["trade_date"],
                    "open": r["open"], "high": r["high"], "low": r["low"],
                    "close": r["close"], "volume": r["volume"],
                } for r in rows]
                n = upsert_macro_batch(db, vals)
                db.commit()
                saved += n
                if on_progress:
                    on_progress(idx, total, f"[FRED {j}/{len(fred_list)}] {series['name']}: {n}件処理")

        # ── 日銀 収集（認証不要・常時）──────────────────────────────────────────
        boj_base_i = len(macro_list) + (len(fred_list) if FRED_API_KEY else 0)
        for k, series in enumerate(boj_list, 1):
            idx = boj_base_i + k
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            if on_progress:
                on_progress(idx - 1, total, f"[BOJ {k}/{len(boj_list)}] {series['name']} 取得中")
            rows = await fetch_boj_series(
                session,
                series["db"],
                series["boj_code"],
                d1_boj,
                d2_boj,
                series.get("lag_days", 0),
                series.get("freq", "monthly"),
            )
            await asyncio.sleep(BOJ_RATE_SLEEP)

            if not rows:
                if on_progress:
                    on_progress(idx, total, f"[BOJ {k}/{len(boj_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total, f"[BOJ {k}/{len(boj_list)}] {series['name']}: {n}件処理")

        # ── OECD 収集（認証不要・常時）────────────────────────────────────────
        oecd_base_i = boj_base_i + len(boj_list)
        for q, series in enumerate(oecd_list, 1):
            idx = oecd_base_i + q
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            if on_progress:
                on_progress(idx - 1, total, f"[OECD {q}/{len(oecd_list)}] {series['name']} 取得中")
            rows = await fetch_oecd_series(
                session,
                series["dataflow"],
                series["series_key"],
                d1_oecd,
                series.get("lag_days", 0),
            )
            await asyncio.sleep(OECD_RATE_SLEEP)

            if not rows:
                if on_progress:
                    on_progress(idx, total, f"[OECD {q}/{len(oecd_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total, f"[OECD {q}/{len(oecd_list)}] {series['name']}: {n}件処理")

        # ── 内閣府ESRI GDP需要項目 収集（認証不要・常時、1回のfetchで4系列取得）──────
        esri_base_i = oecd_base_i + len(oecd_list)
        if cancel_check and cancel_check():
            if on_progress:
                on_progress(esri_base_i, total, "[マクロ収集] ユーザー停止")
            return saved

        if esri_list and on_progress:
            on_progress(esri_base_i, total, f"[ESRI 1/{len(esri_list)}] GDP需要項目CSV 取得中")
        # 1 リクエストで 4 系列ぶんを取る＝ only で全系列が外れたら fetch ごと省く。
        esri_cache = await fetch_esri_gdp_csv(session) if esri_list else {}
        for r_i, series in enumerate(esri_list, 1):
            idx = esri_base_i + r_i
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            base_rows = esri_cache.get(series["esri_column"], [])
            rows = _esri_apply_lag(base_rows, series.get("lag_days", 0))

            if not rows:
                if on_progress:
                    on_progress(idx, total, f"[ESRI {r_i}/{len(esri_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total, f"[ESRI {r_i}/{len(esri_list)}] {series['name']}: {n}件処理")

        # ── IMF WEO 見通し 収集（認証不要・常時、バックフィルは1回のfetchで全系列取得）──
        imf_base_i = esri_base_i + len(esri_list)
        if cancel_check and cancel_check():
            if on_progress:
                on_progress(imf_base_i, total, "[マクロ収集] ユーザー停止")
            return saved

        if imf_list and on_progress:
            on_progress(imf_base_i, total, f"[IMF 1/{len(imf_list)}] WEO historical 取得中")
        # ESRI と同型：バックフィルは 1 回の fetch で全系列ぶん取るので only で空なら省く。
        imf_hist_cache = await fetch_imf_weo_historical(session) if imf_list else {}
        for h_i, series in enumerate(imf_list, 1):
            idx = imf_base_i + h_i
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            hist_rows = imf_hist_cache.get(series["code"], [])
            current_rows = await fetch_imf_weo_current(session, series["indicator"])
            rows = hist_rows + current_rows

            if not rows:
                if on_progress:
                    on_progress(idx, total, f"[IMF {h_i}/{len(imf_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total, f"[IMF {h_i}/{len(imf_list)}] {series['name']}: {n}件処理")

        # ── e-Stat 収集（ESTAT_API_KEY が設定されている場合のみ）────────────────
        if not ESTAT_API_KEY:
            if on_progress:
                on_progress(total, total, "[e-Stat] ESTAT_API_KEY 未設定のためスキップ")
        else:
            estat_base_i = imf_base_i + len(imf_list)
            for m, series in enumerate(estat_list, 1):
                idx = estat_base_i + m
                if cancel_check and cancel_check():
                    if on_progress:
                        on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                    return saved

                if on_progress:
                    on_progress(idx - 1, total, f"[e-Stat {m}/{len(estat_list)}] {series['name']} 取得中")
                rows = await fetch_estat_series(
                    session,
                    series["stats_data_id"],
                    series["cd_tab"],
                    series["cd_cat01"],
                    series["cd_area"],
                    d1_estat,
                    d2_estat,
                    series.get("lag_days", 0),
                )

                if not rows:
                    if on_progress:
                        on_progress(idx, total, f"[e-Stat {m}/{len(estat_list)}] {series['name']} データ無し")
                    continue

                vals = [{
                    "series_code": series["code"],
                    "series_name": series["name"],
                    "category":    series["category"],
                    "trade_date":  r["trade_date"],
                    "open": r["open"], "high": r["high"], "low": r["low"],
                    "close": r["close"], "volume": r["volume"],
                } for r in rows]
                n = upsert_macro_batch(db, vals)
                db.commit()
                saved += n
                if on_progress:
                    on_progress(idx, total, f"[e-Stat {m}/{len(estat_list)}] {series['name']}: {n}件処理")

            # ── e-Stat 鉱工業指数（time 軸が連番コード・日付範囲パラメータ無し）────
            estat_idx_base_i = estat_base_i + len(estat_list)
            for p, series in enumerate(estat_idx_list, 1):
                idx = estat_idx_base_i + p
                if cancel_check and cancel_check():
                    if on_progress:
                        on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                    return saved

                if on_progress:
                    on_progress(idx - 1, total, f"[e-Stat-idx {p}/{len(estat_idx_list)}] {series['name']} 取得中")
                rows = await fetch_estat_index_series(
                    session,
                    series["stats_data_id"],
                    series["cd_cat01"],
                    series.get("lag_days", 0),
                )

                if not rows:
                    if on_progress:
                        on_progress(idx, total, f"[e-Stat-idx {p}/{len(estat_idx_list)}] {series['name']} データ無し")
                    continue

                vals = [{
                    "series_code": series["code"],
                    "series_name": series["name"],
                    "category":    series["category"],
                    "trade_date":  r["trade_date"],
                    "open": r["open"], "high": r["high"], "low": r["low"],
                    "close": r["close"], "volume": r["volume"],
                } for r in rows]
                n = upsert_macro_batch(db, vals)
                db.commit()
                saved += n
                if on_progress:
                    on_progress(idx, total, f"[e-Stat-idx {p}/{len(estat_idx_list)}] {series['name']}: {n}件処理")

        # ── GDELT 収集（認証不要・常時・#406）───────────────────────────────────
        # 1系列＝1リクエストで全履歴が返る。レート制限（1req/5s）は fetch 側でリトライ、
        # 系列間にも GDELT_RATE_SLEEP を挟む。
        gdelt_base_i = (
            imf_base_i + len(imf_list)
            + (len(estat_list) + len(estat_idx_list) if ESTAT_API_KEY else 0)
        )
        for g, series in enumerate(gdelt_list, 1):
            idx = gdelt_base_i + g
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            if on_progress:
                on_progress(idx - 1, total, f"[GDELT {g}/{len(gdelt_list)}] {series['name']} 取得中")
            rows = await fetch_gdelt_timeline(
                session, series["query"], series["mode"], d1_gdelt, d2_gdelt
            )
            await asyncio.sleep(GDELT_RATE_SLEEP)

            if not rows:
                if on_progress:
                    on_progress(idx, total, f"[GDELT {g}/{len(gdelt_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total, f"[GDELT {g}/{len(gdelt_list)}] {series['name']}: {n}件処理")

        # ── Wikimedia Pageviews 収集（認証不要・常時・#406）──────────────────────
        wiki_base_i = gdelt_base_i + len(gdelt_list)
        for w, series in enumerate(wiki_list, 1):
            idx = wiki_base_i + w
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            if on_progress:
                on_progress(idx - 1, total,
                            f"[Wikimedia {w}/{len(wiki_list)}] {series['name']} 取得中")
            rows = await fetch_wikimedia_pageviews(
                session, series["articles"], d1_wiki, d2_wiki
            )
            await asyncio.sleep(WIKIMEDIA_RATE_SLEEP)

            if not rows:
                if on_progress:
                    on_progress(idx, total,
                                f"[Wikimedia {w}/{len(wiki_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total,
                            f"[Wikimedia {w}/{len(wiki_list)}] {series['name']}: {n}件処理")

        # ── 財務省 国債金利 収集（認証不要・常時・#458）────────────────────────
        # 全期間版は月次更新で当月分を含まないため、**初回だけ全期間版で埋め、以降は
        # 当月版で差分**を取る。判定は「DB の最古 trade_date が要求窓の起点より新しいか」
        # ＝years_back を伸ばしたときも自然に埋め直しが走る（件数0固定の判定にしない）。
        mof_base_i = wiki_base_i + len(wiki_list)
        for f_i, series in enumerate(mof_list, 1):
            idx = mof_base_i + f_i
            if cancel_check and cancel_check():
                if on_progress:
                    on_progress(idx - 1, total, "[マクロ収集] ユーザー停止")
                return saved

            existing_min = (
                db.query(sqla_func.min(MacroData.trade_date))
                .filter(MacroData.series_code == series["code"]).scalar()
            )
            need_full = existing_min is None or existing_min > start.isoformat()
            if on_progress:
                on_progress(idx - 1, total,
                            f"[MOF {f_i}/{len(mof_list)}] {series['name']} "
                            f"{'全期間版' if need_full else '当月版'} 取得中")
            # 全期間版は**月次更新で当月分を含まない**ため、初回は当月版も併せて取る
            # （そうしないと初回収集直後だけ最大1か月ぶん古いまま＝鮮度ゲートが鳴る）。
            rows = await fetch_mof_jgb_csv(session, series["mof_column"], full=need_full)
            if need_full:
                rows = rows + await fetch_mof_jgb_csv(
                    session, series["mof_column"], full=False)
            # 基準日は観測日そのもの。公表は翌営業日なので lag_days 分だけ後ろへ寄せて
            # 「この日には知れた値」へ正規化する（ESRI と同じ扱い・ADR-0028）。
            rows = _esri_apply_lag(rows, series.get("lag_days", 0))

            if not rows:
                if on_progress:
                    on_progress(idx, total,
                                f"[MOF {f_i}/{len(mof_list)}] {series['name']} データ無し")
                continue

            vals = [{
                "series_code": series["code"],
                "series_name": series["name"],
                "category":    series["category"],
                "trade_date":  r["trade_date"],
                "open": r["open"], "high": r["high"], "low": r["low"],
                "close": r["close"], "volume": r["volume"],
            } for r in rows]
            n = upsert_macro_batch(db, vals)
            db.commit()
            saved += n
            if on_progress:
                on_progress(idx, total,
                            f"[MOF {f_i}/{len(mof_list)}] {series['name']}: {n}件処理")

    return saved
