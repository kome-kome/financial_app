"""update_industry_from_jpx のユニットテスト (#78)。

HTTP 呼び出しをモックし、Company/FinancialRecord の業種が
バルク UPDATE で正しく更新されることを検証する。
"""
import asyncio
import io
import os
import sys

import httpx
import openpyxl
import pytest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from collector import _read_jpx_excel, update_industry_from_jpx


def _make_jpx_xlsx(rows: list) -> bytes:
    """テスト用 JPX 業種マスタ xlsx を生成（col1=sec_code, col5=industry）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["col0", "sec_code", "col2", "col3", "col4", "industry_name"])
    for sec, ind in rows:
        ws.append(["", sec, "", "", "", ind])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mock_client(content: bytes) -> httpx.AsyncClient:
    def handler(request):
        return httpx.Response(200, content=content)
    return httpx.AsyncClient(transport=httpx.MockTransport(handler))


class _FakeXlsSheet:
    """xlrd の sheet インターフェース（cell_value / nrows）を最小再現するスタブ。"""
    def __init__(self, rows):
        self._rows = rows
    @property
    def nrows(self):
        return len(self._rows)
    def cell_value(self, row, col):
        return self._rows[row][col]   # 列不足は IndexError（実 xlrd と同挙動）


class _FakeXlsBook:
    def __init__(self, rows):
        self._sheet = _FakeXlsSheet(rows)
    def sheet_by_index(self, idx):
        return self._sheet


class TestReadJpxExcel:
    """Excel バイト列 → 業種辞書の純粋変換 `_read_jpx_excel` の単体テスト。"""

    def test_parses_real_xlsx(self):
        """.xlsx 形式（xlrd が XLRDError → openpyxl フォールバック経路）を実データで検証。"""
        content = _make_jpx_xlsx([("1001", "情報・通信業"), ("2001", "小売業")])
        result = _read_jpx_excel(content)
        assert result == {"1001": "情報・通信業", "2001": "小売業"}

    def test_xlsx_skips_blank_industry_rows(self):
        """業種が空（'-'/None）の行はスキップされる。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["col0", "sec_code", "c2", "c3", "c4", "industry"])
        ws.append(["", "2001", "", "", "", "小売業"])   # 採用
        ws.append(["", "3001", "", "", "", "-"])        # 業種 '-' → スキップ
        ws.append(["", "4001", "", "", "", None])       # 業種 None → スキップ
        buf = io.BytesIO(); wb.save(buf)
        result = _read_jpx_excel(buf.getvalue())
        assert result == {"2001": "小売業"}

    def test_empty_sheet_returns_empty_dict(self):
        """ヘッダのみ（データ行なし）の空シートは空辞書を返す。"""
        wb = openpyxl.Workbook()
        wb.active.append(["col0", "sec_code", "c2", "c3", "c4", "industry"])
        buf = io.BytesIO(); wb.save(buf)
        assert _read_jpx_excel(buf.getvalue()) == {}

    def test_missing_required_columns_skipped(self):
        """業種列（6列目）を欠く行は IndexError をスキップして空辞書になる。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["col0", "sec_code"])      # 2列しかない
        ws.append(["", "1001"])              # 業種列なし → スキップ
        buf = io.BytesIO(); wb.save(buf)
        assert _read_jpx_excel(buf.getvalue()) == {}

    def test_parses_xls_via_xlrd(self, monkeypatch):
        """.xls 形式（xlrd 成功経路）を、xlrd.open_workbook をスタブ化して検証。"""
        rows = [
            ["col0", "sec_code", "c2", "c3", "c4", "industry"],  # ヘッダ
            ["", 1001.0, "", "", "", "情報・通信業"],            # xls は数値が float
            ["", 101.0, "", "", "", "建設業"],                   # float → 4桁ゼロ埋め "0101"
            ["", "9999", "", "", "", "サービス業"],
        ]
        import xlrd
        monkeypatch.setattr(xlrd, "open_workbook",
                            lambda *a, **k: _FakeXlsBook(rows))
        result = _read_jpx_excel(b"dummy-xls-bytes")
        assert result == {"1001": "情報・通信業", "0101": "建設業", "9999": "サービス業"}


class TestUpdateIndustryFromJpx:
    def _run(self, coro):
        return asyncio.run(coro)

    def test_updates_company_industry(self, db, make_company):
        db.add(make_company(edinet_code="E00001", sec_code="1001", industry=""))
        db.commit()
        content = _make_jpx_xlsx([("1001", "情報・通信業")])
        client = _mock_client(content)
        co_updated, fr_updated = self._run(update_industry_from_jpx(client, db))
        assert co_updated == 1
        from database import Company
        co = db.query(Company).filter_by(edinet_code="E00001").first()
        assert co.industry == "情報・通信業"

    def test_updates_financial_record_industry(self, db, make_company, make_fin):
        db.add(make_company(edinet_code="E00001", sec_code="2001"))
        db.add(make_fin(edinet_code="E00001", sec_code="2001", industry="旧業種"))
        db.commit()
        content = _make_jpx_xlsx([("2001", "小売業")])
        client = _mock_client(content)
        _, fr_updated = self._run(update_industry_from_jpx(client, db))
        assert fr_updated == 1
        from database import FinancialRecord
        fr = db.query(FinancialRecord).filter_by(edinet_code="E00001").first()
        assert fr.industry == "小売業"

    def test_no_change_when_industry_already_correct(self, db, make_company):
        db.add(make_company(edinet_code="E00001", sec_code="1001"))
        db.query(type(make_company())).filter_by(edinet_code="E00001").update({"industry": "情報・通信業"})
        db.commit()
        content = _make_jpx_xlsx([("1001", "情報・通信業")])
        client = _mock_client(content)
        co_updated, _ = self._run(update_industry_from_jpx(client, db))
        assert co_updated == 0

    def test_zero_padded_sec_code_matches(self, db, make_company):
        # DB に "0101"、JPX マップは "0101" → ゼロ埋め形式での一致
        db.add(make_company(edinet_code="E00001", sec_code="0101"))
        db.commit()
        content = _make_jpx_xlsx([("0101", "建設業")])
        client = _mock_client(content)
        co_updated, _ = self._run(update_industry_from_jpx(client, db))
        assert co_updated == 1

    def test_http_error_returns_zeros(self, db):
        def handler(request):
            return httpx.Response(500)
        client = httpx.AsyncClient(transport=httpx.MockTransport(handler))
        co_updated, fr_updated = self._run(update_industry_from_jpx(client, db))
        assert co_updated == 0
        assert fr_updated == 0
